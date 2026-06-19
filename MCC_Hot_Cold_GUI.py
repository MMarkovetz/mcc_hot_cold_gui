"""
MCC_Hot_Cold_GUI - Python port of the MATLAB App Designer application.

Deposition and MCC Analyzer. Widget names match the MATLAB classdef so
callback bodies can be ported from extracted_matlab_code.m one function
at a time. Image loading + display + radio-toggle are wired up; the
analysis callbacks (Blur, Threshold, Mask, Analyze, MCC, Export) are
stubbed with TODOs pointing to their MATLAB source line numbers.
"""

from __future__ import annotations

import os
import sys
from pathlib import Path

# ---------------------------------------------------------------------------
# Isolate PySide6's Qt DLLs from any other Qt libraries on PATH.
# MUST run before `from PySide6...` - see README for details.
# ---------------------------------------------------------------------------
if sys.platform == "win32":
    _bad_keywords = ("matlab", "anaconda", "miniconda",
                     "qgis", "kicad", "\\qt\\", "\\qt5\\", "\\qt6\\")
    _clean = [d for d in os.environ.get("PATH", "").split(os.pathsep)
              if not any(k in d.lower() for k in _bad_keywords)]
    os.environ["PATH"] = os.pathsep.join(_clean)
    try:
        import PySide6  # noqa: F401
        _pyside_dir = Path(PySide6.__file__).resolve().parent
        if hasattr(os, "add_dll_directory"):
            os.add_dll_directory(str(_pyside_dir))
        _plugins = _pyside_dir / "plugins" / "platforms"
        if _plugins.is_dir():
            os.environ.setdefault("QT_QPA_PLATFORM_PLUGIN_PATH", str(_plugins))
    except ImportError:
        pass

import numpy as np
from PySide6.QtCore import Qt
from PySide6.QtGui import QFont
from PySide6.QtWidgets import (
    QApplication, QButtonGroup, QCheckBox, QComboBox, QFileDialog,
    QGroupBox, QLabel, QLineEdit, QMainWindow, QMessageBox, QPushButton,
    QRadioButton, QSlider, QSpinBox, QWidget,
)
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure

try:
    import pydicom
except ImportError:
    pydicom = None

try:
    import scipy.io as sio
except ImportError:
    sio = None

try:
    import scipy.ndimage as ndi
    import scipy.stats as sstats
except ImportError:
    ndi = None
    sstats = None

# Module-level list to keep popup figure windows from being garbage-collected
# (Qt windows that lose their last Python reference get destroyed silently).
_popup_windows = []


# Base figure size, in App Designer's native 1x logical pixels.
BASE_W, BASE_H = 660, 505

# UI scale factor. >1 makes everything bigger. Set MCC_UI_SCALE in env
# to override (e.g. set MCC_UI_SCALE=1.6).
UI_SCALE = float(os.environ.get("MCC_UI_SCALE", "1.30"))


def _s(v):
    """Scale a single value by UI_SCALE."""
    return int(round(v * UI_SCALE))


def mpos(x, y, w, h, parent_h=BASE_H):
    """MATLAB (bottom-left) base-pixel position -> scaled Qt (top-left) geometry."""
    return _s(x), _s(parent_h - y - h), _s(w), _s(h)


def gpos(x, y, w, h):
    """Group-box internal child position (already in MATLAB-up coords),
    pre-flipped by the caller; just scale."""
    return _s(x), _s(y), _s(w), _s(h)


class UIAxesCanvas(FigureCanvas):
    """matplotlib canvas standing in for the MATLAB uiaxes."""

    def __init__(self, parent=None):
        self._figure = Figure(figsize=(4, 4), tight_layout=True)
        super().__init__(self._figure)
        if parent is not None:
            self.setParent(parent)
        self.ax = self._figure.add_subplot(111)
        self.ax.set_xticks([]); self.ax.set_yticks([])
        self._blank = np.full((256, 256), 255, dtype=np.uint8)
        self.ax.imshow(self._blank, cmap="gray", vmin=0, vmax=255)
        self.SmoothTx = None
        self.CurrentIm = None
        self.draw_idle()

    def show_image(self, arr, cmap="gray"):
        self.ax.clear()
        self.ax.set_xticks([]); self.ax.set_yticks([])
        self.ax.imshow(arr, cmap=cmap)
        self.CurrentIm = arr
        self.draw_idle()

    def clear(self):
        self.show_image(self._blank)


class FigureWindow(QMainWindow):
    """Standalone matplotlib figure window (replaces MATLAB `figure`).

    Each call shows a separate, resizable, savable plot.  Keeps a strong
    reference in the module-level _popup_windows list so Qt does not
    garbage-collect it the instant `AnalyzeDepoButtonPushed` returns.
    """

    def __init__(self, title="Figure", size=(700, 600)):
        super().__init__()
        self.setWindowTitle(title)
        self.resize(*size)
        self.figure = Figure(tight_layout=True)
        self.canvas = FigureCanvas(self.figure)
        try:
            from matplotlib.backends.backend_qtagg import (
                NavigationToolbar2QT as NavToolbar)
            self._toolbar = NavToolbar(self.canvas, self)
            self.addToolBar(self._toolbar)
        except Exception:
            self._toolbar = None
        self.setCentralWidget(self.canvas)
        self.ax = self.figure.add_subplot(111)
        _popup_windows.append(self)
        # Drop the reference when the user closes the window.
        self.destroyed.connect(lambda *_: _popup_windows.remove(self)
                               if self in _popup_windows else None)


class DraggablePolygon:
    """Interactive polygon overlay for a matplotlib axes.

    Features
    --------
    * Click+drag any vertex to move it.
    * Click+drag inside the polygon to translate the whole shape.
    * Right-click a vertex to delete it (minimum 3 vertices kept).
    * Double-click an edge to insert a new vertex at the click point.

    Notes
    -----
    The polygon is drawn as two artists - a closed Line2D for the edges
    and a scatter (Line2D with markers only) for the vertex handles -
    so that picking a vertex is a simple distance check in display
    coordinates and works regardless of axis scaling / aspect ratio.
    """

    PICK_RADIUS_PX = 10  # vertex pick radius in screen pixels
    EDGE_PICK_PX   = 8   # how close (in px) a double-click must be to an edge
    EDGE_COLOR     = (0.0, 0.447, 0.741)
    HANDLE_COLOR   = (1.0, 0.55, 0.0)

    def __init__(self, canvas, vertices, on_change=None):
        self.canvas = canvas        # UIAxesCanvas (FigureCanvas)
        self.ax = canvas.ax
        self.on_change = on_change
        # Ensure a closed list of vertices internally; we re-close on draw.
        v = np.asarray(vertices, dtype=float)
        if v.shape[0] >= 2 and np.allclose(v[0], v[-1]):
            v = v[:-1]              # store as OPEN ring
        self.vertices = v.copy()

        # Edge line (closed ring rendered explicitly).
        from matplotlib.lines import Line2D
        ring = np.vstack([self.vertices, self.vertices[0:1]])
        self.line = Line2D(ring[:, 0], ring[:, 1],
                           color=self.EDGE_COLOR, linewidth=2,
                           solid_capstyle="round")
        self.ax.add_line(self.line)

        # Vertex handles (scatter via Line2D markers - cheap to update).
        self.handles = Line2D(self.vertices[:, 0], self.vertices[:, 1],
                              linestyle="None", marker="o",
                              markersize=7,
                              markerfacecolor=self.HANDLE_COLOR,
                              markeredgecolor="black",
                              markeredgewidth=1, zorder=10)
        self.ax.add_line(self.handles)

        # Drag state.
        self._drag_idx = None       # vertex index being dragged
        self._drag_offset = None    # for whole-shape translation: (dx, dy) at press
        self._drag_mode = None      # 'vertex' | 'translate' | None

        # Connect events.
        self._cids = [
            canvas.mpl_connect("button_press_event",   self._on_press),
            canvas.mpl_connect("motion_notify_event",  self._on_motion),
            canvas.mpl_connect("button_release_event", self._on_release),
        ]

    # -- public ------------------------------------------------------
    def disconnect(self):
        for cid in self._cids:
            try:
                self.canvas.mpl_disconnect(cid)
            except Exception:
                pass
        self._cids = []
        for art in (self.line, self.handles):
            try:
                art.remove()
            except Exception:
                pass

    # -- helpers -----------------------------------------------------
    def _emit_change(self):
        self._refresh_artists()
        self.canvas.draw_idle()
        if self.on_change is not None:
            try:
                self.on_change(self.vertices)
            except Exception:
                import traceback; traceback.print_exc()

    def _refresh_artists(self):
        ring = np.vstack([self.vertices, self.vertices[0:1]])
        self.line.set_data(ring[:, 0], ring[:, 1])
        self.handles.set_data(self.vertices[:, 0], self.vertices[:, 1])

    def _vertex_at(self, event):
        """Return the index of the vertex under the cursor, or None."""
        if event.x is None or event.y is None:
            return None
        # Convert all vertices to display coords and measure distances.
        disp = self.ax.transData.transform(self.vertices)
        d = np.hypot(disp[:, 0] - event.x, disp[:, 1] - event.y)
        idx = int(np.argmin(d))
        return idx if d[idx] <= self.PICK_RADIUS_PX else None

    def _edge_hit(self, event):
        """Return (i, point) where i is the index of the edge endpoint
        BEFORE the insertion, and `point` is the projection on that edge,
        or None if no edge is within EDGE_PICK_PX."""
        if event.x is None or event.y is None:
            return None
        disp = self.ax.transData.transform(self.vertices)
        n = len(disp)
        best = None
        best_d = self.EDGE_PICK_PX
        for i in range(n):
            p1 = disp[i]
            p2 = disp[(i + 1) % n]
            seg = p2 - p1
            seg_len2 = np.dot(seg, seg)
            if seg_len2 == 0:
                continue
            t = np.dot(np.array([event.x, event.y]) - p1, seg) / seg_len2
            t = max(0.0, min(1.0, t))
            proj = p1 + t * seg
            d = float(np.hypot(proj[0] - event.x, proj[1] - event.y))
            if d < best_d:
                best_d = d
                # Convert insertion point back to data coords.
                inv = self.ax.transData.inverted().transform(proj)
                best = (i, inv)
        return best

    def _inside(self, x, y):
        """Even-odd point-in-polygon test in data coords."""
        v = self.vertices
        n = len(v)
        inside = False
        j = n - 1
        for i in range(n):
            xi, yi = v[i]
            xj, yj = v[j]
            if ((yi > y) != (yj > y)) and \
               (x < (xj - xi) * (y - yi) / (yj - yi + 1e-30) + xi):
                inside = not inside
            j = i
        return inside

    # -- event handlers ---------------------------------------------
    def _on_press(self, event):
        if event.inaxes is not self.ax:
            return
        # Right-click: delete the vertex under the cursor (min 3 verts).
        if event.button == 3:
            idx = self._vertex_at(event)
            if idx is not None and len(self.vertices) > 3:
                self.vertices = np.delete(self.vertices, idx, axis=0)
                self._emit_change()
            return
        # Double-click on an edge: insert a vertex.
        if event.dblclick and event.button == 1:
            hit = self._edge_hit(event)
            if hit is not None:
                i, pt = hit
                self.vertices = np.insert(self.vertices, i + 1, pt, axis=0)
                self._emit_change()
                return
        # Plain left click.
        if event.button == 1:
            idx = self._vertex_at(event)
            if idx is not None:
                self._drag_mode = "vertex"
                self._drag_idx = idx
                return
            if event.xdata is not None and self._inside(event.xdata, event.ydata):
                self._drag_mode = "translate"
                self._drag_offset = (event.xdata, event.ydata)

    def _on_motion(self, event):
        if event.inaxes is not self.ax or event.xdata is None:
            return
        if self._drag_mode == "vertex" and self._drag_idx is not None:
            self.vertices[self._drag_idx] = (event.xdata, event.ydata)
            self._emit_change()
        elif self._drag_mode == "translate" and self._drag_offset is not None:
            dx = event.xdata - self._drag_offset[0]
            dy = event.ydata - self._drag_offset[1]
            self.vertices = self.vertices + np.array([dx, dy])
            self._drag_offset = (event.xdata, event.ydata)
            self._emit_change()

    def _on_release(self, event):
        self._drag_mode = None
        self._drag_idx = None
        self._drag_offset = None


class ManualROIDialog(QMainWindow):
    """Modal-ish popup that lets the user adjust the polygon for one MCC
    frame.  Blocks via a Qt event loop until the user clicks "Next Frame"
    (commit) or "Cancel" (abort entire manual run).

    Returns the new Nx2 vertex array via .result_vertices, or None if the
    user cancelled.
    """

    def __init__(self, frame_2d, init_vertices, frame_index, n_frames,
                 parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Manual ROI - frame {frame_index + 1} / {n_frames}")
        self.resize(640, 720)
        from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout,
                                       QPushButton, QLabel)

        self.result_vertices = None
        self._cancelled = False

        central = QWidget(self); self.setCentralWidget(central)
        v = QVBoxLayout(central)

        # Canvas with the frame image.
        self._figure = Figure(tight_layout=True)
        self._canvas = FigureCanvas(self._figure)
        self._ax = self._figure.add_subplot(111)
        self._ax.imshow(np.asarray(frame_2d, dtype=float), cmap="gray")
        self._ax.set_xticks([]); self._ax.set_yticks([])
        v.addWidget(self._canvas, 1)

        # Wrap the matplotlib canvas in a tiny shim so DraggablePolygon
        # (which expects a UIAxesCanvas-like object with .ax) is happy.
        class _Shim:
            def __init__(s, canvas, ax):
                s.canvas = canvas; s.ax = ax
            def mpl_connect(s, *a, **kw):
                return s.canvas.mpl_connect(*a, **kw)
            def mpl_disconnect(s, *a, **kw):
                return s.canvas.mpl_disconnect(*a, **kw)
            def draw_idle(s):
                s.canvas.draw_idle()
        shim = _Shim(self._canvas, self._ax)
        self._editor = DraggablePolygon(shim, init_vertices)

        # Hint label.
        v.addWidget(QLabel(
            "Drag any vertex to reshape.  Drag inside to translate.  "
            "Right-click a vertex to delete.  Double-click an edge to insert."))

        # Buttons.
        row = QHBoxLayout()
        next_btn = QPushButton("Next Frame ▶")
        next_btn.setDefault(True)
        next_btn.clicked.connect(self._accept)
        cancel_btn = QPushButton("Cancel manual run")
        cancel_btn.clicked.connect(self._reject)
        row.addWidget(cancel_btn)
        row.addStretch(1)
        row.addWidget(next_btn)
        v.addLayout(row)

        self._canvas.draw_idle()

    def _accept(self):
        self.result_vertices = self._editor.vertices.copy()
        self.close()

    def _reject(self):
        self._cancelled = True
        self.result_vertices = None
        self.close()

    def exec_modal(self):
        """Show the window and pump events until it closes."""
        from PySide6.QtCore import QEventLoop
        loop = QEventLoop()
        self.destroyed.connect(loop.quit)
        # Also break the loop if the window simply closes (no destroyed yet).
        self._loop_killer = loop
        self.show()
        loop.exec()
        return self.result_vertices, self._cancelled

    def closeEvent(self, event):
        # Make sure the event loop wakes up.
        try:
            self._loop_killer.quit()
        except Exception:
            pass
        super().closeEvent(event)


class ScanStackViewer(QMainWindow):
    """Popup viewer that steps through every frame of an MCC scan stack.

    Optional am_transforms (one per frame i>=1, mapping ref->frame i) and
    am_centroids (a list of [(x,y),(x,y)] per frame) drive a moving ROI
    overlay and fiducial markers when Am-241 tracking was used.

    Equivalent to the MATLAB GetMCCButton "ShowScan" block (line 971):
    the user advances frame-by-frame via Prev/Next buttons, a slider, or
    the left/right arrow keys.  The polygon ROI is drawn on top of every
    frame.
    """

    def __init__(self, stack, polygon_xy, bkg_median=0.0, times=None,
                 cmap="viridis", parent=None,
                 am_transforms=None, am_centroids=None):
        super().__init__(parent)
        self.setWindowTitle("MCC Scan Stack")
        self.resize(720, 720)
        from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout,
                                       QPushButton, QSlider, QLabel)

        self._stack = stack
        self._poly  = np.asarray(polygon_xy, dtype=float)
        self._bkg   = float(bkg_median)
        self._n     = int(stack.shape[0])
        self._am_transforms = am_transforms        # list of 2x3 forward Ms
        self._am_centroids  = am_centroids         # list of [(x,y),(x,y)] per frame
        self._times = (np.asarray(times) if times is not None
                       else np.arange(self._n))
        self._cmap  = cmap

        central = QWidget(self); self.setCentralWidget(central)
        v = QVBoxLayout(central)

        self._figure = Figure(tight_layout=True)
        self._canvas = FigureCanvas(self._figure)
        self._ax = self._figure.add_subplot(111)
        v.addWidget(self._canvas, 1)

        ctrl = QHBoxLayout()
        self._prev_btn  = QPushButton("◀ Prev"); self._prev_btn.clicked.connect(self.prev_frame)
        self._next_btn  = QPushButton("Next ▶"); self._next_btn.clicked.connect(self.next_frame)
        self._slider    = QSlider(Qt.Horizontal)
        self._slider.setRange(0, max(0, self._n - 1))
        self._slider.valueChanged.connect(self.show_frame)
        self._label     = QLabel("Frame 0 / 0")
        ctrl.addWidget(self._prev_btn)
        ctrl.addWidget(self._slider, 1)
        ctrl.addWidget(self._next_btn)
        ctrl.addWidget(self._label)
        v.addLayout(ctrl)

        self._idx = 0
        if self._n > 0:
            self.show_frame(0)

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Left:
            self.prev_frame()
        elif event.key() == Qt.Key_Right:
            self.next_frame()
        else:
            super().keyPressEvent(event)

    def prev_frame(self):
        if self._idx > 0:
            self._slider.setValue(self._idx - 1)

    def next_frame(self):
        if self._idx < self._n - 1:
            self._slider.setValue(self._idx + 1)

    def show_frame(self, idx):
        self._idx = int(idx)
        frame = self._stack[self._idx]
        if frame.ndim == 3 and frame.shape[-1] == 1:
            frame = frame[..., 0]
        elif frame.ndim == 3:
            frame = frame.mean(axis=-1)
        img = np.asarray(frame, dtype=float) - self._bkg

        self._ax.clear()
        self._ax.imshow(img, cmap=self._cmap)

        # Move the polygon by this frame's Am-241 transform, if available.
        poly = self._poly
        if (self._am_transforms is not None and self._idx > 0
                and self._idx - 1 < len(self._am_transforms)):
            M = self._am_transforms[self._idx - 1]
            # Apply M @ [x, y, 1]^T to every vertex.
            ones = np.ones((poly.shape[0], 1))
            homog = np.hstack([poly, ones])
            poly  = (M @ homog.T).T
        ring = np.vstack([poly, poly[0:1]])
        self._ax.plot(ring[:, 0], ring[:, 1],
                      color=(0, 0.447, 0.741), linewidth=2)

        # Show the detected fiducial centroids for this frame, if known.
        if (self._am_centroids is not None
                and self._idx < len(self._am_centroids)):
            for (cx, cy) in self._am_centroids[self._idx]:
                self._ax.plot(cx, cy, "o", markersize=12,
                              markerfacecolor="none",
                              markeredgecolor=(0, 1, 0),
                              markeredgewidth=2)
        self._ax.set_xticks([]); self._ax.set_yticks([])
        t = self._times[self._idx] if self._idx < len(self._times) else self._idx
        title_t = f"t = {t:g} min" if isinstance(t, (int, float, np.floating)) else f"Frame {self._idx}"
        self._ax.set_title(f"{title_t}  ({self._idx + 1} / {self._n})")
        self._label.setText(f"Frame {self._idx + 1} / {self._n}")
        self._canvas.draw_idle()


class MCCHotColdGUI(QMainWindow):
    """Python port of MCC_Hot_Cold_GUI.mlapp."""

    def _init_state(self):
        """Mirrors the MATLAB `properties (Access = private)` block."""
        self.HotColdIm = None
        self.BKGim = None
        self.TxIm = None
        self.DepoIM = None
        self.MCCstack = None
        self.ROIposition = None
        self.RLmask = None
        self.BKGimpath = ""
        self.Tximpath = ""
        self.MCCstackpath = ""
        self.SubDir = ""
        self.RawTx = None
        self.SmoothTx = None
        self.imWasModified = False
        self.maskPersist = None
        self.HCdone = False
        self.MCCdone = False
        self.HotPix = None
        self.ColdPix = None
        self.HotColdData = None
        self.MCCarray = None
        self.FastSlowArray = None
        self.HotMask = None
        self.ColdMask = None
        self.CustomROI = False
        self.FastSlowArray90 = None
        self.AUC60 = None
        self.AUC90 = None
        self.maskedDepoIm = None
        self.MCCcmap = None
        self.OG_mask_pos = None
        self.OG_mask_moved = False
        self.Tx_Translation = None
        self.Cmask = None
        self.Pmask = None

    def __init__(self):
        super().__init__()
        self._init_state()
        self.setWindowTitle("Deposition and MCC Analyzer")
        self.setMinimumSize(_s(BASE_W), _s(BASE_H))
        self.resize(_s(BASE_W), _s(BASE_H))
        self._central = QWidget(self)
        self.setCentralWidget(self._central)
        self._create_components()
        self._startup_fcn()

    # ------------------------------------------------------------------ UI
    def _create_components(self):
        p = self._central

        # UIAxes (image canvas)
        self.UIAxes = UIAxesCanvas(p)
        self.UIAxes.setGeometry(*mpos(14, 141, 376, 345))

        # BKG / Tx / Scan Stack text fields + labels + Browse buttons
        self.BKGImageEditFieldLabel = self._label(
            p, "BKG Image", mpos(1, 84, 68, 22),
            align=Qt.AlignRight | Qt.AlignVCenter)
        self.BKGImageEditField = QLineEdit(p)
        self.BKGImageEditField.setGeometry(*mpos(84, 84, 100, 22))
        self.BKGImageEditField.editingFinished.connect(self.BKGImageEditFieldValueChanged)
        self.BKGImageEditField.textEdited.connect(self.BKGImageEditFieldValueChanging)

        self.TxImageEditFieldLabel = self._label(
            p, "Tx Image", mpos(1, 52, 68, 22),
            align=Qt.AlignRight | Qt.AlignVCenter)
        self.TxImageEditField = QLineEdit(p)
        self.TxImageEditField.setGeometry(*mpos(84, 52, 100, 22))
        self.TxImageEditField.editingFinished.connect(self.TxImageEditFieldValueChanged)
        self.TxImageEditField.textEdited.connect(self.TxImageEditFieldValueChanging)

        self.ScanStackEditFieldLabel = self._label(
            p, "Scan Stack", mpos(3, 20, 66, 22),
            align=Qt.AlignRight | Qt.AlignVCenter)
        self.ScanStackEditField = QLineEdit(p)
        self.ScanStackEditField.setGeometry(*mpos(84, 20, 100, 22))
        self.ScanStackEditField.editingFinished.connect(self.ScanStackEditFieldValueChanged)
        self.ScanStackEditField.textEdited.connect(self.ScanStackEditFieldValueChanging)

        self.Browse_for_BKG = QPushButton("Browse", p)
        self.Browse_for_BKG.setGeometry(*mpos(193, 85, 100, 22))
        self.Browse_for_BKG.clicked.connect(self.Browse_for_BKGButtonPushed)

        self.Browse_for_Tx = QPushButton("Browse", p)
        self.Browse_for_Tx.setGeometry(*mpos(193, 52, 100, 22))
        self.Browse_for_Tx.clicked.connect(self.Browse_for_TxButtonPushed)

        self.Browse_for_Scan = QPushButton("Browse", p)
        self.Browse_for_Scan.setGeometry(*mpos(193, 20, 100, 22))
        self.Browse_for_Scan.clicked.connect(self.Browse_for_ScanButtonPushed)

        # Display radio group
        self.DisplayButtonGroup = QGroupBox("Display", p)
        self.DisplayButtonGroup.setGeometry(*mpos(300, 14, 105, 99))
        gh = 99  # base group-box height
        self.BKGButton = QRadioButton("BKG", self.DisplayButtonGroup)
        self.BKGButton.setGeometry(*gpos(11, gh - 53 - 22, 58, 22))
        self.BKGButton.setChecked(True)
        self.TxButton = QRadioButton("Tx", self.DisplayButtonGroup)
        self.TxButton.setGeometry(*gpos(11, gh - 31 - 22, 65, 22))
        self.DepositionButton = QRadioButton("Deposition", self.DisplayButtonGroup)
        self.DepositionButton.setGeometry(*gpos(11, gh - 9 - 22, 79, 22))
        self._display_group = QButtonGroup(self)
        for b in (self.BKGButton, self.TxButton, self.DepositionButton):
            self._display_group.addButton(b)
        self._display_group.buttonClicked.connect(self.DisplayButtonGroupSelectionChanged)

        # Section headers
        self.ImportPatientImagesLabel = self._label(
            p, "Import Patient Images", mpos(166, 112, 152, 22),
            align=Qt.AlignCenter, bold=True, size=14)
        self.ProcessScanLabel = self._label(
            p, "Process Scan", mpos(485, 483, 97, 22),
            align=Qt.AlignCenter, bold=True, size=14)

        # ------------------------------------------------------------------
        # Sections 1-3 (Filter / Threshold / Make Masks) are temporarily
        # disabled. The user's normal workflow uses the default adjustable
        # polygon mask (Section 4) instead. Re-enable any of these blocks
        # by un-commenting them.
        # ------------------------------------------------------------------
        # # 1. Filter (Width)
        # self.FilterWidthLabel = self._label(
        #     p, "1. Filter (Width)", mpos(399, 452, 93, 22),
        #     align=Qt.AlignRight | Qt.AlignVCenter, bold=True)
        # self.FilterWidthEditField = QSpinBox(p)
        # self.FilterWidthEditField.setRange(0, 999999)
        # self.FilterWidthEditField.setValue(1)
        # self.FilterWidthEditField.setGeometry(*mpos(499, 452, 34, 22))
        # self.FilterWidthEditField.valueChanged.connect(self.FilterWidthEditFieldValueChanged)
        # self.BlurButton = QPushButton("Blur", p)
        # self.BlurButton.setGeometry(*mpos(544, 452, 48, 22))
        # self.BlurButton.clicked.connect(self.BlurButtonPushed)
        # self.UndoBlurButton = QPushButton("Undo", p)
        # self.UndoBlurButton.setGeometry(*mpos(599, 452, 48, 22))
        # self.UndoBlurButton.clicked.connect(self.UndoBlurButtonPushed)
        #
        # # 2. Threshold
        # self.ThresholdSliderLabel = self._label(
        #     p, "2. Threshold", mpos(399, 402, 77, 22),
        #     align=Qt.AlignRight | Qt.AlignVCenter, bold=True)
        # self.ThresholdSlider = QSlider(Qt.Horizontal, p)
        # self.ThresholdSlider.setRange(0, 1000)
        # self.ThresholdSlider.setGeometry(*mpos(499, 402, 137, 22))
        # self.ThresholdSlider.valueChanged.connect(
        #     lambda v: self.ThresholdSliderValueChanging(v / 1000.0))
        #
        # # 3. Make Masks
        # self.MakeMasksLabel = self._label(
        #     p, "3. Make Masks", mpos(404, 343, 89, 22), bold=True)
        # self.DilateandFillButton = QPushButton("Dilate and Fill", p)
        # self.DilateandFillButton.setGeometry(*mpos(499, 343, 93, 22))
        # self.DilateandFillButton.clicked.connect(self.DilateandFillButtonPushed)
        # self.GetsButton = QPushButton("Get #s", p)
        # self.GetsButton.setGeometry(*mpos(594, 343, 57, 22))
        # self.GetsButton.clicked.connect(self.GetsButtonPushed)

        # 1. Mask RL  (was section 4 in the original MATLAB app; sections 1-3
        # are commented out above. Shifted up by 163 px to fill the gap.)
        self.MaskRLDropDownLabel = self._label(
            p, "1. Mask RL", mpos(399, 452, 68, 22),
            align=Qt.AlignRight | Qt.AlignVCenter, bold=True)
        self.MaskRLDropDown = QComboBox(p)
        self.MaskRLDropDown.setGeometry(*mpos(499, 452, 43, 22))
        self.MaskRLDropDown.addItems(["Default", "From File"])
        self.MaskButton = QPushButton("Mask", p)
        self.MaskButton.setGeometry(*mpos(544, 452, 48, 22))
        self.MaskButton.clicked.connect(self.MaskButtonPushed)
        self.UnmaskButton = QPushButton("Unmask", p)
        self.UnmaskButton.setGeometry(*mpos(595, 452, 57, 22))
        self.UnmaskButton.clicked.connect(self.UnmaskButtonPushed)

        # Lung ROI file + Track
        self.LungROIFileEditFieldLabel = self._label(
            p, "Lung ROI File", mpos(404, 423, 80, 22),
            align=Qt.AlignRight | Qt.AlignVCenter)
        self.LungROIFileEditField = QLineEdit(p)
        self.LungROIFileEditField.setGeometry(*mpos(499, 423, 100, 22))
        self.LungROIFileEditField.editingFinished.connect(self.LungROIFileEditFieldValueChanged)
        self.LungROIFileEditField.textEdited.connect(self.LungROIFileEditFieldValueChanging)
        self.TrackCheckBox = QCheckBox("Track", p)
        self.TrackCheckBox.setGeometry(*mpos(604, 424, 52, 22))

        # Helper hint about the adjustable polygon
        self.PolygonHintLabel = self._label(
            p, "Tip: drag any vertex to reshape the mask",
            mpos(404, 395, 252, 18),
            align=Qt.AlignLeft | Qt.AlignVCenter)
        f = QFont(self.PolygonHintLabel.font())
        f.setItalic(True); f.setPointSize(max(8, f.pointSize() - 1))
        self.PolygonHintLabel.setFont(f)
        self.PolygonHintLabel.setStyleSheet("color: gray;")

        # 2. Get Hot/Cold  (was section 5)
        self.GetHotColdLabel = self._label(
            p, "2. Get Hot/Cold", mpos(404, 360, 92, 22), bold=True)
        self.AnalyzeDepoButton = QPushButton("Analyze Depo", p)
        self.AnalyzeDepoButton.setGeometry(*mpos(499, 360, 92, 22))
        self.AnalyzeDepoButton.clicked.connect(self.AnalyzeDepoButtonPushed)

        # 3. Get WL MCC  (was section 6)
        self.GetWLMCCLabel = self._label(
            p, "3. Get WL MCC", mpos(404, 312, 91, 22), bold=True)
        self.Am241CheckBox = QCheckBox("Am-241", p)
        self.Am241CheckBox.setGeometry(*mpos(499, 312, 64, 22))
        self.Am241CheckBox.stateChanged.connect(self.Am241CheckBoxValueChanged)
        self.ShowScanCheckBox = QCheckBox("Show Scan", p)
        self.ShowScanCheckBox.setGeometry(*mpos(573, 312, 83, 22))
        self.ShowScanCheckBox.stateChanged.connect(self.ShowScanCheckBoxValueChanged)
        self.GetMCCButton = QPushButton("Get MCC", p)
        self.GetMCCButton.setGeometry(*mpos(499, 291, 70, 22))
        self.GetMCCButton.clicked.connect(self.GetMCCButtonPushed)
        self.ManuallyCheckBox = QCheckBox("Manually", p)
        self.ManuallyCheckBox.setGeometry(*mpos(573, 291, 70, 22))
        self.ManuallyCheckBox.stateChanged.connect(self.ManuallyCheckBoxValueChanged)

        # Export / Reset
        self.ExportDataButton = QPushButton("Export Data", p)
        ef = self.ExportDataButton.font(); ef.setBold(True)
        self.ExportDataButton.setFont(ef)
        self.ExportDataButton.setGeometry(*mpos(485, 105, 97, 22))
        self.ExportDataButton.clicked.connect(self.ExportDataButtonPushed)
        self.ResetNewScanButton = QPushButton("Reset/New Scan", p)
        self.ResetNewScanButton.setGeometry(*mpos(547, 20, 105, 22))
        self.ResetNewScanButton.clicked.connect(self.ResetNewScanButtonPushed)

    @staticmethod
    def _label(parent, text, geom, *,
               align=Qt.AlignLeft | Qt.AlignVCenter,
               bold=False, size=None):
        lbl = QLabel(text, parent)
        lbl.setGeometry(*geom)
        lbl.setAlignment(align)
        if bold or size is not None:
            f = QFont(lbl.font())
            if bold:
                f.setBold(True)
            if size is not None:
                f.setPointSize(size)
            lbl.setFont(f)
        return lbl

    # =================================================================
    # Callbacks
    # =================================================================

    def _startup_fcn(self):
        """TODO: MATLAB line 448. Port startupFcn."""
        self.UIAxes.clear()

    # ---- DICOM reading & image display -------------------------------
    @staticmethod
    def _read_dicom(path):
        """Read a DICOM file. Returns (arr_4d, n_frames, samples_per_pixel).

        arr_4d is always shaped (frames, H, W, channels) so the rest of the
        code can index uniformly. Channels is 1 for grayscale, 3 for RGB.
        """
        if pydicom is None:
            raise RuntimeError(
                "pydicom is not installed.  pip install pydicom")
        ds = pydicom.dcmread(path, force=True)
        try:
            arr = ds.pixel_array
        except Exception as e:
            # Most common cause: compressed Transfer Syntax without decoder.
            ts = getattr(ds.file_meta, "TransferSyntaxUID", "?")
            raise RuntimeError(
                f"pydicom could not decode pixel data (TransferSyntaxUID={ts}).\n"
                f"If the data is JPEG/JPEG2000/RLE-compressed, install:\n"
                f"  pip install pylibjpeg pylibjpeg-libjpeg pylibjpeg-openjpeg "
                f"python-gdcm\n\nOriginal error: {e}") from e

        n = int(getattr(ds, "NumberOfFrames", 1) or 1)
        samples = int(getattr(ds, "SamplesPerPixel", 1) or 1)

        # Reshape to (frames, H, W, channels) regardless of pydicom's choice.
        a = np.asarray(arr)
        if samples == 1:
            if a.ndim == 2:               # (H, W)
                a = a[None, :, :, None]
            elif a.ndim == 3:             # (F, H, W)
                a = a[:, :, :, None]
            elif a.ndim == 4:             # already (F, H, W, 1) or similar
                if a.shape[-1] != 1:
                    a = a[..., None]
            else:
                raise RuntimeError(f"Unexpected grayscale DICOM shape {a.shape}")
        else:  # RGB / multi-channel
            if a.ndim == 3 and a.shape[-1] in (3, 4):     # (H, W, C)
                a = a[None, ...]
            elif a.ndim == 4:                              # (F, H, W, C)
                pass
            else:
                raise RuntimeError(f"Unexpected color DICOM shape {a.shape}")
        return a, n, samples

    @staticmethod
    def _frame_to_2d(frame):
        """Collapse a single (H, W, C) frame to a 2-D array for grayscale display."""
        if frame.ndim == 2:
            return frame
        if frame.ndim == 3:
            if frame.shape[-1] == 1:
                return frame[..., 0]
            return frame.mean(axis=-1)            # luminance
        return frame

    def updateimage(self, imagefile):
        """Port of MATLAB updateimage() (extracted_matlab_code.m:91-188)."""
        try:
            im = None
            if isinstance(imagefile, np.ndarray):
                im = imagefile
            elif imagefile in ("", None):
                im = np.full((256, 256), 255, dtype=np.uint8)
            elif (isinstance(imagefile, (str, Path))
                  and str(imagefile).lower().endswith(".dcm")):
                arr4d, n, _samples = self._read_dicom(str(imagefile))
                # arr4d shape: (frames, H, W, channels)
                if self.BKGButton.isChecked():
                    idx = 0 if n == 1 else 1
                    im = self._frame_to_2d(arr4d[idx])
                    self.BKGim = im
                elif self.TxButton.isChecked():
                    idx = 0 if n == 1 else 1
                    im = self._frame_to_2d(arr4d[idx])
                    # TODO: port imtranslate (scipy.ndimage.shift)
                    self.TxIm = im
                elif self.DepositionButton.isChecked():
                    if n == 1:
                        idx = 0
                    elif n == 47:
                        idx = 0
                    else:
                        idx = min(45, n - 1)
                    im = self._frame_to_2d(arr4d[idx])
                    self.DepoIM = im
                    if n > 47:
                        self.MCCstack = arr4d[45:90]
                    else:
                        self.MCCstack = arr4d
            else:
                import matplotlib.image as mpimg
                im = mpimg.imread(str(imagefile))

            if im is None:
                return

            # Matplotlib is happiest with float arrays; clip outliers.
            disp = np.asarray(im)
            if disp.dtype.kind in ("u", "i"):
                disp = disp.astype(np.float32)
            self.UIAxes.show_image(disp)

        except Exception as e:
            import traceback
            QMessageBox.critical(
                self, "Image error",
                f"{e}\n\n{traceback.format_exc()}")
    # ---- Browse + Display ---------------------------------------------
    def _start_dir(self):
        return self.SubDir or ""

    def Browse_for_BKGButtonPushed(self):
        """MATLAB line 486."""
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Background Image", self._start_dir(),
            "DICOM files (*.dcm);;All files (*)")
        if not path:
            return
        if not self.SubDir:
            self.SubDir = str(Path(path).parent)
        self.BKGimpath = path
        self.BKGImageEditField.setText(Path(path).name)
        self.BKGButton.setChecked(True)
        self.updateimage(path)

    def Browse_for_TxButtonPushed(self):
        """MATLAB line 542."""
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Transmission Image", self._start_dir(),
            "DICOM / TIFF (*.dcm *.tif *.tiff);;All files (*)")
        if not path:
            return
        if not self.SubDir:
            self.SubDir = str(Path(path).parent)
        self.Tximpath = path
        self.TxImageEditField.setText(Path(path).name)
        self.TxButton.setChecked(True)
        self.updateimage(path)

    def Browse_for_ScanButtonPushed(self):
        """MATLAB line 583."""
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Scan Stack", self._start_dir(),
            "DICOM files (*.dcm);;All files (*)")
        if not path:
            return
        if not self.SubDir:
            self.SubDir = str(Path(path).parent)
        self.MCCstackpath = path
        self.ScanStackEditField.setText(Path(path).name)
        self.DepositionButton.setChecked(True)
        self.updateimage(path)

    def DisplayButtonGroupSelectionChanged(self, *_):
        """MATLAB line 519. Switch canvas between BKG / Tx / Deposition."""
        if self.BKGButton.isChecked() and self.BKGimpath:
            self.updateimage(self.BKGimpath)
        elif self.TxButton.isChecked() and self.Tximpath:
            self.updateimage(self.Tximpath)
        elif self.DepositionButton.isChecked() and self.MCCstackpath:
            self.updateimage(self.MCCstackpath)

    def ResetNewScanButtonPushed(self):
        """MATLAB line 1161. Reset all state."""
        if (QMessageBox.question(
                self, "Reset", "Reset all state and start a new scan?",
                QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes):
            self._init_state()
            self.UIAxes.clear()
            for le in (self.BKGImageEditField, self.TxImageEditField,
                       self.ScanStackEditField, self.LungROIFileEditField):
                le.clear()

    # ---- Mask -------------------------------------------------------
    # Hardcoded default lung-shaped polygon (15 vertices), in 128x128 pixels.
    # Will be scaled to the loaded image size (per MATLAB MaskButtonPushed).
    _DEFAULT_LUNG_ROI_128 = np.array([
        [70, 42], [71, 55], [72, 71], [73, 82], [83, 82], [95, 82],
        [107, 81], [104, 66], [103, 55], [101, 40], [91, 21], [80, 18],
        [73, 19], [70, 30], [70, 42],
    ], dtype=float)

    def _draw_polygon_overlay(self, vertices):
        """Draw an *adjustable* polygon overlay on the canvas.

        vertices: Nx2 array of (x, y).  Click+drag any vertex to reshape
        the mask.  Right-click a vertex to delete it.  Double-click on an
        edge to insert a new vertex at that point.  self.ROIposition is
        kept in sync so downstream callbacks (AnalyzeDepo, Track, ...)
        always read the current mask shape.
        """
        # Remove any prior overlay first.
        self._remove_polygon_overlay()
        self._poly_editor = DraggablePolygon(
            self.UIAxes, np.asarray(vertices, dtype=float),
            on_change=self._on_polygon_changed,
        )
        # Mirror DraggablePolygon -> self for back-compat reads.
        self._roi_artist = self._poly_editor.line
        self.ROIposition = self._poly_editor.vertices.copy()
        self.UIAxes.draw_idle()

    def _remove_polygon_overlay(self):
        editor = getattr(self, "_poly_editor", None)
        if editor is not None:
            editor.disconnect()
        self._poly_editor = None
        self._roi_artist = None

    def _on_polygon_changed(self, new_vertices):
        """Called by DraggablePolygon whenever the user reshapes the mask."""
        self.ROIposition = np.asarray(new_vertices, dtype=float).copy()
        # Tracking flag - mirror MATLAB AnalyzeDepoButton logic.
        og = getattr(self, "OG_mask_pos", None)
        if og is not None and og.shape == self.ROIposition.shape:
            if not np.allclose(og, self.ROIposition):
                self.OG_mask_moved = True
                self.Tx_Translation = self.ROIposition - og

    def _load_roi_from_mat(self, mat_path):
        r"""Return Nx2 polygon vertices from a .mat file.

        The MATLAB app saves an *entire* images.roi.Polygon handle object,
        not just its Position field:

            lung_mask = app.ROIposition;
            save([app.SubDir,'\patient_roi.mat'], 'lung_mask')

        scipy.io.loadmat can't fully introspect MATLAB handle classes, so the
        recovered `lung_mask` may be a struct, an mat_struct, an opaque
        object array, or a record array. This routine tries hard to find an
        Nx2 array of vertices anywhere inside.
        """
        if sio is None:
            raise RuntimeError("scipy is required to load .mat files. "
                               "Run: pip install scipy")

        # ---- Try the classic v7 loader first. ------------------------
        data = None
        try:
            data = sio.loadmat(mat_path, squeeze_me=True,
                               struct_as_record=False)
        except NotImplementedError as e:
            # v7.3 .mat files are HDF5 - need h5py.
            try:
                import h5py
            except ImportError:
                raise RuntimeError(
                    "This is a MATLAB v7.3 (HDF5) .mat file. "
                    "Install h5py:  pip install h5py\n\n"
                    f"Original error: {e}") from e
            return self._load_roi_from_hdf5(mat_path)

        # ---- Walk every variable looking for vertices ---------------
        keys = [k for k in data.keys() if not k.startswith("__")]
        found = self._find_vertex_array(data, keys)
        if found is not None:
            return found

        # ---- Fallback: scan MatlabOpaque byte buffers ---------------
        # The MATLAB app saves the entire images.roi.Polygon object;
        # scipy returns it as a MatlabOpaque whose 'arr' field is the
        # serialized object data. Position is in there as raw doubles.
        scanned = self._scan_opaque_for_vertices(data)
        if scanned is not None:
            return scanned

        # ---- Diagnostics if we still couldn't find anything ----------
        summary = []
        for k in keys:
            v = data[k]
            summary.append(f"  '{k}' -> {type(v).__name__}, "
                           f"shape={getattr(v, 'shape', 'n/a')}, "
                           f"dtype={getattr(v, 'dtype', 'n/a')}")
        raise RuntimeError(
            "Could not find an Nx2 vertex array in the .mat file.\n\n"
            "The MATLAB app saves the entire images.roi.Polygon HANDLE "
            "object, which scipy cannot always introspect. The file's "
            "top-level variables are:\n"
            + "\n".join(summary)
            + "\n\nWorkaround: in MATLAB, re-export just the vertex array:\n"
              "    load patient_roi.mat            % loads lung_mask\n"
              "    vertices = lung_mask.Position;  % Nx2 double\n"
              "    save patient_roi_pos.mat vertices\n"
              "...and try loading that file instead.")

    def _find_vertex_array(self, obj, hint_keys=None, depth=0):
        """Recursively search a scipy-loaded .mat structure for an Nx2 array."""
        if depth > 6:                       # safety brake
            return None

        # 1) Direct ndarray match.
        if isinstance(obj, np.ndarray):
            # Plain numeric Nx2 - this is what we want.
            if (obj.ndim == 2 and obj.shape[1] == 2 and
                    obj.dtype.kind in "fiub" and obj.shape[0] >= 3):
                return obj.astype(float)
            # 2xN is also acceptable - just transpose.
            if (obj.ndim == 2 and obj.shape[0] == 2 and
                    obj.dtype.kind in "fiub" and obj.shape[1] >= 3):
                return obj.T.astype(float)
            # Object / structured array - recurse into each entry.
            if obj.dtype == object or obj.dtype.names is not None:
                if obj.dtype.names:
                    for name in obj.dtype.names:
                        v = self._find_vertex_array(obj[name], depth=depth + 1)
                        if v is not None:
                            return v
                for item in obj.ravel():
                    v = self._find_vertex_array(item, depth=depth + 1)
                    if v is not None:
                        return v

        # 2) A scipy mat_struct (struct_as_record=False).
        if hasattr(obj, "_fieldnames"):
            # Prefer the 'Position' field if it exists.
            for pref in ("Position", "position", "Vertices", "vertices"):
                if pref in obj._fieldnames:
                    v = self._find_vertex_array(getattr(obj, pref),
                                                depth=depth + 1)
                    if v is not None:
                        return v
            for name in obj._fieldnames:
                v = self._find_vertex_array(getattr(obj, name),
                                            depth=depth + 1)
                if v is not None:
                    return v

        # 3) Dict (top-level loadmat result).
        if isinstance(obj, dict):
            # Visit preferred names first.
            ordered = []
            for pref in ("lung_mask", "Position", "vertices", "roi"):
                if pref in obj:
                    ordered.append(pref)
            ordered.extend(k for k in obj if k not in ordered
                           and not k.startswith("__"))
            for k in ordered:
                v = self._find_vertex_array(obj[k], depth=depth + 1)
                if v is not None:
                    return v

        return None

    @staticmethod
    def _scan_opaque_for_vertices(data,
                                   min_vertices=5, max_vertices=400,
                                   min_val=-32.0, max_val=4096.0,
                                   min_spread=5.0):
        """Hunt for an Nx2 polygon inside any MatlabOpaque byte buffer.

        MATLAB serializes images.roi.Polygon as an MCOS class whose
        'Position' property is an Nx2 double array. scipy doesn't
        decode MCOS, but the raw bytes survive in the 'arr' field.
        We scan every uint8 buffer reachable from `data` for the
        longest run of doubles that:
        - are all finite,
        - fall in a plausible pixel-coord range,
        - have non-trivial spatial spread,
        - decode to an even-length array (=> Nx2).
        """
        import struct, math

        buffers = []

        def _collect(obj, depth=0):
            if depth > 8:
                return
            if isinstance(obj, dict):
                for v in obj.values():
                    _collect(v, depth + 1)
                return
            if isinstance(obj, np.ndarray):
                if obj.dtype == np.uint8:
                    buffers.append(obj.tobytes())
                    return
                if obj.dtype.names:
                    for name in obj.dtype.names:
                        _collect(obj[name], depth + 1)
                if obj.dtype == object:
                    for item in obj.ravel():
                        _collect(item, depth + 1)
            elif hasattr(obj, "_fieldnames"):
                for name in obj._fieldnames:
                    _collect(getattr(obj, name), depth + 1)
            elif isinstance(obj, (bytes, bytearray)):
                buffers.append(bytes(obj))

        _collect(data)
        if not buffers:
            return None

        # Two tracks: among "good" candidates (score <= SCORE_OK) prefer the
        # longest; fall back to lowest score if none qualify. Real polygons
        # have many vertices AND low scores; random byte coincidences usually
        # have one or the other but not both.
        SCORE_OK = 0.5
        good_best = None; good_n = 0; good_s = float("inf")
        any_best  = None; any_s   = float("inf")

        for buf in buffers:
            n = len(buf)
            if n < min_vertices * 16:
                continue
            # Sweep 8-byte aligned offsets.
            off = 0
            while off + 16 <= n:
                # Walk forward 8 bytes at a time as long as the doubles
                # stay finite and inside [min_val, max_val].
                run = []
                cur = off
                while cur + 8 <= n:
                    (d,) = struct.unpack_from("<d", buf, cur)
                    if not math.isfinite(d):
                        break
                    if not (min_val <= d <= max_val):
                        break
                    # Reject denormalized/sub-pixel junk: real polygon
                    # coords are >= ~0.5 pixel or exactly 0.
                    if d != 0.0 and abs(d) < 0.1:
                        break
                    run.append(d)
                    cur += 8
                    if len(run) >= max_vertices * 2:
                        break
                # Sweep multiple truncation lengths of this run and try BOTH
                # reshape conventions.  MATLAB stores arrays in COLUMN-MAJOR
                # (Fortran) order, so an Nx2 Position is laid out
                # [x1..xN, y1..yN] - reshape(2,-1).T.  An array saved by
                # Python (or any C-order tool) is row-major [x1,y1, x2,y2,...]
                # - reshape(-1, 2).  We score each candidate on:
                #   - both columns must have non-trivial spread (kills the
                #     "all zeros in one column" failure when the run extends
                #     into MCOS padding bytes),
                #   - max consecutive step relative to bounding-box diagonal
                #     (real polygons have ordered vertices => small steps),
                #   - first<->last vertex closure (MATLAB ROIs close).
                # The (offset, length, reshape) tuple with the smallest score
                # wins.
                run_len = len(run)
                if run_len >= min_vertices * 2:
                    flat_full = np.array(run)
                    # Sweep even lengths from longest down to 2*min_vertices,
                    # in steps of 2 (each step removes one row from arr_row
                    # or one column from arr_col).
                    L = run_len - (run_len % 2)
                    while L >= min_vertices * 2:
                        flat = flat_full[:L]
                        for arr in (flat.reshape(-1, 2),
                                    flat.reshape(2, -1).T):
                            score = MCCHotColdGUI._polygon_quality(arr, min_spread)
                            if score == float("inf"):
                                continue
                            n_v = arr.shape[0]
                            if score <= SCORE_OK:
                                if (n_v > good_n or
                                        (n_v == good_n and score < good_s)):
                                    good_best = arr.copy()
                                    good_n = n_v
                                    good_s = score
                            if score < any_s:
                                any_best = arr.copy()
                                any_s = score
                        L -= 2
                off += 8  # always advance, never skip ahead

        return good_best if good_best is not None else any_best

    @staticmethod
    def _polygon_quality(arr, min_spread=5.0):
        """Score an Nx2 polygon candidate. Lower is better; inf means reject.

        A real polygon has:
          - both columns with non-trivial spread (not all zeros / not
            constant),
          - vertices ordered around the perimeter so consecutive steps are
            small relative to the bounding-box diagonal,
          - first vertex close to last (closure).
        """
        if arr.shape[0] < 4:
            return float("inf")
        xs, ys = arr[:, 0], arr[:, 1]
        x_range = float(xs.max() - xs.min())
        y_range = float(ys.max() - ys.min())
        # Reject if either coordinate has near-zero spread (the "all y=0"
        # failure when MCOS padding zeros leak into one reshape column).
        if x_range < min_spread or y_range < min_spread:
            return float("inf")
        if float(xs.std()) < min_spread / 4 or float(ys.std()) < min_spread / 4:
            return float("inf")
        # Reject runs that contain extended (x, y) == (0, 0) padding.
        zero_rows = int(np.sum((np.abs(xs) < 1e-9) & (np.abs(ys) < 1e-9)))
        if zero_rows > 1:
            return float("inf")
        diag = float(np.hypot(x_range, y_range))
        steps = np.linalg.norm(np.diff(arr, axis=0), axis=1)
        max_step = float(steps.max())
        # If the biggest consecutive step is huge relative to the bbox,
        # the vertices are not ordered around the perimeter.
        if max_step > diag:
            return float("inf")
        closure = float(np.linalg.norm(arr[-1] - arr[0]))
        # Score: prefer small max-step and good closure, normalized by bbox.
        return (max_step + 2.0 * closure) / (diag + 1e-9)

    def _load_roi_from_hdf5(self, mat_path):
        """Fallback for MATLAB v7.3 (HDF5) .mat files. Requires h5py."""
        import h5py
        with h5py.File(mat_path, "r") as f:
            # Walk every dataset; pick the first one shaped Nx2 (or 2xN).
            picked = []

            def _visit(name, node):
                if isinstance(node, h5py.Dataset):
                    arr = node[()]
                    if isinstance(arr, np.ndarray) and arr.ndim == 2:
                        if arr.shape[1] == 2 and arr.shape[0] >= 3:
                            picked.append(arr.astype(float))
                        elif arr.shape[0] == 2 and arr.shape[1] >= 3:
                            picked.append(arr.T.astype(float))

            f.visititems(_visit)
            if picked:
                return picked[0]
        raise RuntimeError(
            "No Nx2 vertex array found inside the HDF5 .mat file.")

    def MaskButtonPushed(self):                        # MATLAB 686
        """Show a polygon ROI overlay on the canvas based on the dropdown."""
        # Switch to the Deposition image first (per MATLAB).
        self.DepositionButton.setChecked(True)
        if self.MCCstackpath:
            self.updateimage(self.MCCstackpath)

        choice = self.MaskRLDropDown.currentText()
        try:
            if choice == "Default":
                # Scale 128x128 polygon to current image size.
                im = self.UIAxes.CurrentIm
                if im is None:
                    QMessageBox.warning(self, "No image",
                        "Load a deposition scan before applying a mask.")
                    return
                imh = im.shape[0] if im.ndim >= 2 else 128
                verts = self._DEFAULT_LUNG_ROI_128 * (imh / 128.0)
                self._draw_polygon_overlay(verts)
                self.OG_mask_pos = verts.copy()
                self.maskPersist = True

            elif choice == "From File":
                roi_path = self.LungROIFileEditField.text().strip().strip('"')
                # Auto-prompt if the field is empty OR points to a nonexistent file.
                if not roi_path or not Path(roi_path).is_file():
                    if roi_path and not Path(roi_path).is_file():
                        QMessageBox.information(
                            self, "ROI file not found",
                            f"'{roi_path}' does not exist.\n"
                            f"Pick the saved ROI file instead.")
                    path, _ = QFileDialog.getOpenFileName(
                        self, "Select Lung ROI file", self._start_dir(),
                        "MATLAB ROI files (*.mat);;NumPy (*.npy);;All files (*)")
                    if not path:
                        return
                    roi_path = path
                    self.LungROIFileEditField.setText(roi_path)

                ext = Path(roi_path).suffix.lower()
                if ext == ".npy":
                    verts = np.load(roi_path)
                elif ext == ".mat":
                    verts = self._load_roi_from_mat(roi_path)
                else:
                    raise RuntimeError(
                        f"Unsupported ROI file type '{ext}'. "
                        f"Expected .mat or .npy.")

                verts = np.asarray(verts, dtype=float)
                if verts.ndim != 2 or verts.shape[1] != 2 or verts.shape[0] < 3:
                    raise RuntimeError(
                        f"ROI file produced an array shaped {verts.shape}, "
                        f"but vertices must be Nx2 with N>=3.")

                self._draw_polygon_overlay(verts)
                self.OG_mask_pos = verts.copy()
                self.maskPersist = True

            else:
                # Indexed region (e.g. "1", "2"...) - requires Get #s to have
                # been run, populating self.UIAxes.cstruct. Stub with a notice.
                QMessageBox.information(
                    self, "Not yet ported",
                    "Indexed-region masks (after Get #s) are not yet implemented.\n"
                    "Use 'Default' or 'From File' for now.")
        except Exception as e:
            import traceback
            QMessageBox.critical(self, "Mask error",
                                 f"{e}\n\n{traceback.format_exc()}")

    def UnmaskButtonPushed(self):                      # MATLAB 741
        """Remove the ROI overlay from the canvas."""
        self._remove_polygon_overlay()
        self.UIAxes.draw_idle()
        self.ROIposition = None
        self.RLmask = None
        self.maskPersist = False

    # ---- Hot/Cold analysis -------------------------------------------
    @staticmethod
    def _polygon_to_mask(vertices, shape):
        """createMask equivalent. vertices: Nx2 (x, y); shape: (H, W).

        Returns a boolean 2-D mask of the polygon interior.
        """
        from matplotlib.path import Path as MplPath
        h, w = shape[:2]
        ys, xs = np.mgrid[:h, :w]
        pts = np.column_stack([xs.ravel(), ys.ravel()])
        # Close the polygon if not already closed.
        v = np.asarray(vertices, dtype=float)
        if v.shape[0] >= 2 and not np.allclose(v[0], v[-1]):
            v = np.vstack([v, v[0:1]])
        path = MplPath(v)
        return path.contains_points(pts).reshape(h, w)

    @staticmethod
    def _bbox_from_mask(mask):
        """Return (x, y, w, h) bounding box of nonzero pixels in `mask`."""
        rows = np.any(mask, axis=1)
        cols = np.any(mask, axis=0)
        if not rows.any() or not cols.any():
            return None
        ymin, ymax = np.where(rows)[0][[0, -1]]
        xmin, xmax = np.where(cols)[0][[0, -1]]
        return (float(xmin), float(ymin),
                float(xmax - xmin + 1), float(ymax - ymin + 1))

    @staticmethod
    def _hot_cold_cmap():
        """Reproduce the MATLAB 6-band colormap from the AnalyzeDepoButtonPushed body.

        (0, 0.1) black | [0.1, 0.5) blue | [0.5, 1.0) cyan |
        [1.0, 1.5) green | [1.5, 2.0) yellow | >= 2.0 red.
        """
        from matplotlib.colors import ListedColormap
        rows = (
            [(0, 0, 0)] +
            [(0, 0, 1)] * 4 +
            [(0, 1, 1)] * 5 +
            [(0, 1, 0)] * 5 +
            [(1, 1, 0)] * 5 +
            [(1, 0, 0)] * 5
        )
        return ListedColormap(rows)

    def AnalyzeDepoButtonPushed(self):                 # MATLAB 749
        """Port of MATLAB AnalyzeDepoButtonPushed (extracted_matlab_code.m:749)."""
        try:
            if ndi is None or sstats is None:
                raise RuntimeError(
                    "scipy is required for Analyze Depo. "
                    "Run: pip install scipy")

            if self.ROIposition is None:
                QMessageBox.warning(
                    self, "No mask",
                    "Apply a lung-region mask first (1. Mask RL).")
                return
            if self.DepoIM is None:
                QMessageBox.warning(
                    self, "No deposition image",
                    "Load a Scan Stack before running Analyze Depo.")
                return
            if self.BKGim is None or self.TxIm is None:
                QMessageBox.warning(
                    self, "Missing inputs",
                    "Both BKG and Tx images must be loaded.")
                return

            # ---- Track-checkbox -> translation ------------------------
            if self.TrackCheckBox.isChecked() and self.OG_mask_pos is not None:
                roi_translated = (np.asarray(self.ROIposition)
                                  - np.asarray(self.OG_mask_pos))
                if roi_translated.size:
                    self.Tx_Translation = roi_translated[0, :]
                    self.OG_mask_moved = bool(np.any(self.Tx_Translation != 0))

            if self.OG_mask_moved:
                self.TxButton.setChecked(True)
                if self.Tximpath:
                    self.updateimage(self.Tximpath)

            # ---- Background median ------------------------------------
            bkg_arr = np.asarray(self.BKGim, dtype=float) / 7.5
            with np.errstate(invalid="ignore"):
                col_medians = np.nanmedian(bkg_arr, axis=0)
                bkg_median = float(np.nanmedian(col_medians))

            # ---- Smooth deposition / transmission ---------------------
            depo = ndi.gaussian_filter(
                np.asarray(self.DepoIM, dtype=float) - bkg_median, sigma=2)
            tx = ndi.gaussian_filter(
                np.asarray(self.TxIm, dtype=float) - bkg_median, sigma=2)

            # ---- Right-lung binary mask from polygon ------------------
            mask = self._polygon_to_mask(self.ROIposition, depo.shape)
            self.RLmask = mask
            if not mask.any():
                raise RuntimeError(
                    "Mask is empty. Reposition the polygon over the lung.")

            # ---- Masked deposition figure -----------------------------
            self.maskedDepoIm = depo * mask
            depo_win = FigureWindow("Masked Deposition", size=(640, 600))
            depo_win.ax.imshow(self.maskedDepoIm, cmap="gray")
            depo_win.ax.set_xticks([]); depo_win.ax.set_yticks([])
            depo_win.ax.set_title("Masked Deposition")
            depo_win.canvas.draw_idle()
            depo_win.show()

            # ---- Medians inside the lung ------------------------------
            depo_in = depo[mask]
            tx_in   = tx[mask]
            depo_median = float(np.median(depo_in))
            tx_median   = float(np.median(tx_in))
            if depo_median == 0 or tx_median == 0:
                raise RuntimeError(
                    "Median inside lung mask is zero - cannot normalize.")

            n_depo = depo / depo_median
            n_tx   = tx / tx_median
            with np.errstate(divide="ignore", invalid="ignore"):
                hcim = np.where(n_tx != 0, n_depo / n_tx, 0.0)
            self.HotColdIm = hcim * mask

            # ---- Pixel lists / masks ----------------------------------
            in_lung = self.HotColdIm[mask]
            self.HotPix  = in_lung > 2.0
            self.ColdPix = in_lung < 0.5

            self.HotMask  = self.HotColdIm > 2.0
            self.ColdMask = (self.HotColdIm < 0.5) & mask

            # ---- Number / sum / skew / C/P ----------------------------
            mask_count = float(mask.sum())
            hot_nr  = float(self.HotPix.sum())  / mask_count
            cold_nr = float(self.ColdPix.sum()) / mask_count
            depo_orig = np.asarray(self.DepoIM, dtype=float)
            hot_sr  = (float(depo_orig[self.HotColdIm > 2.0].sum())
                       / float(depo_orig[mask].sum()))
            depo_skew = float(sstats.skew(depo_orig[mask].ravel()))

            # Bounding box -> central / peripheral masks
            bbox = self._bbox_from_mask(mask)
            if bbox is None:
                raise RuntimeError("Could not get bounding box of lung mask.")
            x, y, w, h = bbox
            cbox_verts = np.array([
                [x,           y + 0.25 * h],
                [x + 0.5 * w, y + 0.25 * h],
                [x + 0.5 * w, y + 0.75 * h],
                [x,           y + 0.75 * h],
            ])
            self.Cmask = self._polygon_to_mask(cbox_verts, mask.shape) & mask
            self.Pmask = mask & ~self.Cmask

            ccounts  = float((depo * self.Cmask).sum())
            pcounts  = float((depo * self.Pmask).sum())
            ctcounts = float((tx   * self.Cmask).sum())
            ptcounts = float((tx   * self.Pmask).sum())
            cp_tx    = ctcounts / ptcounts if ptcounts != 0 else np.nan
            cp_ratio = (ccounts / pcounts) / cp_tx \
                if (pcounts != 0 and cp_tx not in (0, np.nan)) else np.nan

            self.HotColdData = np.array(
                [hot_nr, cold_nr, hot_sr, depo_skew, cp_ratio], dtype=float)

            # ---- Hot/Cold map figure ----------------------------------
            hc_win = FigureWindow("Hot/Cold Map", size=(720, 620))
            cmap = self._hot_cold_cmap()
            im = hc_win.ax.imshow(self.HotColdIm, cmap=cmap, vmin=0, vmax=2.5)
            hc_win.ax.set_xticks([]); hc_win.ax.set_yticks([])
            hc_win.ax.set_title("Hot/Cold Map")
            hc_win.figure.colorbar(im, ax=hc_win.ax)
            hc_win.canvas.draw_idle()
            hc_win.show()

            # ---- Stats summary popup ----------------------------------
            stats_msg = (
                f"Hot # ratio       = {hot_nr:.4f}\n"
                f"Cold # ratio      = {cold_nr:.4f}\n"
                f"Hot sum ratio     = {hot_sr:.4f}\n"
                f"Deposition skew   = {depo_skew:.4f}\n"
                f"C/P ratio         = {cp_ratio:.4f}"
            )
            QMessageBox.information(self, "Hot/Cold analysis", stats_msg)

            self.HCdone = True

        except Exception as e:
            import traceback
            QMessageBox.critical(
                self, "Analyze Depo error",
                f"{e}\n\n{traceback.format_exc()}")

    # ---- MCC analysis (Get MCC) -------------------------------------
    def _build_mcc_time_vector(self, n_frames):
        """Mirror MATLAB time-vector logic from GetMCCButtonPushed (line 862).

        Returns a 1-D numpy array of length n_frames giving the imaging time
        (minutes) for each frame.
        """
        scan_name = (self.ScanStackEditField.text() or "").lower()
        if n_frames < 47:
            # 45-frame UNC style: 2, 4, ..., 90.
            return np.arange(2, 91, 2, dtype=float)[:n_frames]
        if "promise" in scan_name:
            # Promise study: paired times then trim last.
            a = np.arange(1, 117, 5, dtype=float)
            b = np.arange(2, 118, 5, dtype=float)
            t = np.empty(a.size + b.size)
            t[0::2] = a; t[1::2] = b
            return t[: n_frames - 1] if n_frames - 1 <= t.size else t
        if self.Am241CheckBox.isChecked():
            return np.arange(0, 93, 2, dtype=float)[:n_frames]
        if n_frames > 90:
            base = list(np.arange(2, 91, 2, dtype=float))
            base += [120.0, 60.0 * 24.0]
            return np.asarray(base, dtype=float)[:n_frames]
        return np.arange(0, 93, 2, dtype=float)[:n_frames]

    @staticmethod
    def _scan_stack_frame_2d(stack, i):
        """Return frame i of MCCstack as a 2-D float array.

        Updateimage stores MCCstack with shape (frames, H, W, channels).
        """
        frame = stack[i]
        if frame.ndim == 3 and frame.shape[-1] == 1:
            frame = frame[..., 0]
        elif frame.ndim == 3:
            frame = frame.mean(axis=-1)
        return np.asarray(frame, dtype=float)

    # ---- Am-241 fiducial tracking -----------------------------------
    @staticmethod
    def _bpass(image, lnoise=2.0, lobject=5, threshold=0.0):
        """Crocker-Grier real-space bandpass filter.

        Port of the bpass() helper at MATLAB line 244.  Convolves the
        image with a Gaussian (lowpass) and a boxcar (the highpass
        partner) and subtracts to get a bandpass.  Edges zeroed to
        suppress convolution artifacts.  Used to locate Am-241 spots.
        """
        from scipy.signal import convolve
        img = np.asarray(image, dtype=float)

        if lnoise == 0:
            gauss_k = np.array([1.0])
        else:
            half = int(np.ceil(5 * lnoise))
            x = np.arange(-half, half + 1)
            gauss_k = np.exp(-(x / (2.0 * lnoise)) ** 2)
            gauss_k = gauss_k / gauss_k.sum()

        if lobject:
            half = int(round(lobject))
            x = np.arange(-half, half + 1)
            box_k = np.ones_like(x, dtype=float)
            box_k = box_k / box_k.sum()

        # Separable 2-D convolution: rows, then columns.
        gconv = convolve(img, gauss_k[None, :], mode="same")
        gconv = convolve(gconv, gauss_k[:, None], mode="same")

        if lobject:
            bconv = convolve(img, box_k[None, :], mode="same")
            bconv = convolve(bconv, box_k[:, None], mode="same")
            filtered = gconv - bconv
        else:
            filtered = gconv

        lzero = int(max(lobject, int(np.ceil(5 * lnoise))))
        if lzero > 0:
            filtered[:lzero, :] = 0
            filtered[-lzero:, :] = 0
            filtered[:, :lzero] = 0
            filtered[:, -lzero:] = 0
        filtered[filtered < threshold] = 0
        return filtered

    @staticmethod
    def _centroids_from_binary(binary_mask):
        """Return list of (x, y) centroids for every connected component."""
        labeled, n = ndi.label(np.asarray(binary_mask) > 0)
        if n == 0:
            return []
        coms = ndi.center_of_mass(np.ones_like(labeled, dtype=float),
                                   labeled, range(1, n + 1))
        # scipy.ndimage returns (row, col) = (y, x); MATLAB regionprops
        # gives (Centroid.x, Centroid.y).  Swap.
        return [(float(c[1]), float(c[0])) for c in coms]

    @staticmethod
    def _similarity_2d_from_pairs(src_pts, dst_pts):
        """Closed-form 2-D similarity transform from at least 2 point pairs.

        Returns a 2x3 forward matrix M such that  dst = M @ [x, y, 1]^T.
        With exactly 2 pairs this reproduces estgeotform2d(...,'similarity').
        """
        src = np.asarray(src_pts, dtype=float)
        dst = np.asarray(dst_pts, dtype=float)
        if src.shape[0] < 2 or dst.shape[0] < 2:
            return np.array([[1.0, 0.0, 0.0],
                             [0.0, 1.0, 0.0]])
        dx = src[1, 0] - src[0, 0]
        dy = src[1, 1] - src[0, 1]
        du = dst[1, 0] - dst[0, 0]
        dv = dst[1, 1] - dst[0, 1]
        denom = dx * dx + dy * dy
        if denom == 0:
            return np.array([[1.0, 0.0, 0.0],
                             [0.0, 1.0, 0.0]])
        # Similarity = complex multiply: (a + ib) * (x + iy) + (tx + i*ty).
        a = (du * dx + dv * dy) / denom
        b = (dv * dx - du * dy) / denom
        tx = dst[0, 0] - (a * src[0, 0] - b * src[0, 1])
        ty = dst[0, 1] - (b * src[0, 0] + a * src[0, 1])
        return np.array([[a, -b, tx],
                         [b,  a, ty]])

    @staticmethod
    def _warp_image(image, M_2x3, output_shape=None, order=1):
        """Forward-affine an image; like MATLAB imwarp + imref2d."""
        if ndi is None:
            return np.asarray(image, dtype=float)
        if output_shape is None:
            output_shape = image.shape[:2]
        Mf = np.eye(3); Mf[:2, :] = M_2x3
        Mi = np.linalg.inv(Mf)
        S = np.array([[0, 1, 0],
                      [1, 0, 0],
                      [0, 0, 1]], dtype=float)
        Mi_rc = S @ Mi @ S
        return ndi.affine_transform(
            np.asarray(image, dtype=float),
            Mi_rc[:2, :2], offset=Mi_rc[:2, 2],
            order=order, mode="constant", cval=0.0,
            output_shape=output_shape)

    @staticmethod
    def _invert_2x3(M):
        Mf = np.eye(3); Mf[:2, :] = M
        Mi = np.linalg.inv(Mf)
        return Mi[:2, :]

    def _get_americium_transforms(self):
        """Port of getAmericiumTransforms (MATLAB line 347).

        Returns (transforms, good_idxs) where:
          transforms[i] is the 2x3 forward similarity transform applied
              to RLmask BEFORE evaluating frame i+1.  Length = n_am-1.
          good_idxs is a list of 1-based frame indices (matching MATLAB)
              where exactly 2 fiducial centroids were detected.
        """
        if pydicom is None:
            raise RuntimeError("pydicom is required for Am-241 tracking.")

        # Look for *clearance_Am.dcm next to the SubDir, else prompt.
        am_path = ""
        if self.SubDir:
            matches = sorted(Path(self.SubDir).glob("*clearance_Am.dcm"))
            if matches:
                am_path = str(matches[0])
        if not am_path:
            am_path, _ = QFileDialog.getOpenFileName(
                self, "Select Am-241 fiducial stack",
                self._start_dir(),
                "DICOM (*.dcm);;All files (*)")
            if not am_path:
                raise RuntimeError(
                    "Am-241 tracking requires a fiducial DICOM stack "
                    "(usually named *clearance_Am.dcm).")

        arr4d, n_frames_am, _samples = self._read_dicom(am_path)
        # arr4d: (frames, H, W, C)
        n_am = int(arr4d.shape[0])

        # ---- Auto-calibrate the threshold on frame 1 ----------------
        # MATLAB hard-codes "bpim > 0.45", which works on their specific
        # DICOM intensity scale.  pydicom returns raw uint16, so the
        # absolute value of 0.45 is meaningless here.  We find the 2
        # reference fiducial peaks on frame 1, then set the threshold to
        # a fraction of the DIMMER peak so it still captures both spots
        # on subsequent frames while rejecting noise / motion fragments.
        first = self._scan_stack_frame_2d(arr4d, 0)
        bp1 = self._bpass(first, 2.0, 5)
        from scipy.ndimage import maximum_filter
        NMS_WIN = 9
        nms = maximum_filter(bp1, size=NMS_WIN)
        peaks = (bp1 == nms) & (bp1 > 0)
        ys, xs = np.where(peaks)
        if len(xs) < 2:
            raise RuntimeError(
                "Bandpass output of frame 1 has fewer than 2 local maxima. "
                "Cannot calibrate Am-241 threshold.")
        intens = bp1[peaks]
        top2 = np.sort(intens)[::-1][:2]
        # Threshold = the configured fraction of the DIMMER reference peak.
        # 0.5 is conservative: captures both spots on frame 1 even if one
        # is dimmer, but still well above noise.
        THRESH_FRAC = 0.5
        threshold = float(top2.min()) * THRESH_FRAC
        self._am_threshold = threshold
        self._am_calibration_peaks = (float(top2[0]), float(top2[1]))

        def _detect(frame_2d):
            """Return centroids of every connected component above threshold.

            Returning ALL connected components is INTENTIONAL: when a
            patient moves during exposure the fiducial smears into >2
            blobs, and we want to reject those frames downstream.
            """
            bp = self._bpass(frame_2d, 2.0, 5)
            return self._centroids_from_binary(bp > threshold)

        # Frame 1 - must have exactly 2 centroids by construction.
        cs = _detect(first)
        if len(cs) != 2:
            raise RuntimeError(
                "Calibration produced {} components on frame 1 (need 2). "
                "Try a different fiducial stack or lower the threshold "
                "fraction.".format(len(cs)))
        # MATLAB y-flip: put the BOTTOM (larger y) point first.
        cs = sorted(cs, key=lambda p: -p[1])
        ctrda = np.array(cs, dtype=float)

        per_frame_centroids = [cs]
        per_frame_n = [2]
        transforms = []
        good_idxs = [1]
        bad_frames = []

        for i in range(1, n_am):
            frame = self._scan_stack_frame_2d(arr4d, i)
            cs = _detect(frame)
            per_frame_n.append(len(cs))
            if len(cs) == 2:
                cs = sorted(cs, key=lambda p: -p[1])
                old_cs = ctrda.copy()
                new = np.array(cs, dtype=float)
                ctrda = new
                M = self._similarity_2d_from_pairs(old_cs, new)
                good_idxs.append(i + 1)
                per_frame_centroids.append(cs)
            else:
                # ANY other count (0, 1, 3+) -> motion or detection
                # failure. Reuse the previous transform, exclude this
                # frame from the retention curve.
                M = self._similarity_2d_from_pairs(ctrda, ctrda)
                per_frame_centroids.append(per_frame_centroids[-1])
                bad_frames.append((i + 1, len(cs)))
            transforms.append(M)

        # Stash everything for the debug popup and the viewer.
        self._am_centroids = per_frame_centroids
        self._am_per_frame_n = per_frame_n
        self._am_first_frame = first
        self._am_bpass_first = bp1
        self._am_bad_frames = bad_frames

        return transforms, good_idxs

    @staticmethod
    def _affine_from_points(src_xy, dst_xy):
        """Fit a 2-D affine transform that maps src points to dst points.

        Returns a 2x3 matrix M such that  [M | [t_x; t_y]] @ [x, y, 1]^T = [x'; y'].
        Solves the least-squares system across all point pairs.
        """
        src = np.asarray(src_xy, dtype=float)
        dst = np.asarray(dst_xy, dtype=float)
        n = min(src.shape[0], dst.shape[0])
        if n < 3:
            return np.array([[1.0, 0.0, 0.0],
                             [0.0, 1.0, 0.0]])
        A = np.zeros((2 * n, 6))
        b = np.zeros(2 * n)
        for i in range(n):
            x, y = src[i]
            xp, yp = dst[i]
            A[2 * i,     :] = [x, y, 1, 0, 0, 0]
            A[2 * i + 1, :] = [0, 0, 0, x, y, 1]
            b[2 * i]     = xp
            b[2 * i + 1] = yp
        params, *_ = np.linalg.lstsq(A, b, rcond=None)
        return params.reshape(2, 3)

    @staticmethod
    def _warp_mask(mask, M_2x3):
        """Apply an affine transform to a binary mask.

        M_2x3 maps src->dst in image-coordinate space (column = x, row = y).
        scipy.ndimage.affine_transform expects the INVERSE mapping
        (output -> input) in matrix-coordinate space (row, col), so we
        invert and swap axes.
        """
        if ndi is None:
            return mask.copy()
        # Build the 3x3 forward matrix.
        Mf = np.eye(3)
        Mf[:2, :] = M_2x3
        Mi = np.linalg.inv(Mf)
        # scipy expects (row, col) ordering. Swap x<->y rows AND columns.
        S = np.array([[0, 1, 0],
                      [1, 0, 0],
                      [0, 0, 1]], dtype=float)
        Mi_rc = S @ Mi @ S
        warped = ndi.affine_transform(
            mask.astype(float), Mi_rc[:2, :2], offset=Mi_rc[:2, 2],
            order=0, mode="constant", cval=0.0,
            output_shape=mask.shape)
        return warped > 0.5


    def _show_am_debug_popup(self):
        """Pop a window showing the Am-241 frame 1 detection so the user
        can verify the algorithm is finding the right spots.

        Also shows the per-frame fiducial count so motion-blurred frames
        (count != 2) are visible at a glance.
        """
        if (not hasattr(self, "_am_first_frame")
                or self._am_first_frame is None):
            return
        frame = self._am_first_frame
        bp    = getattr(self, "_am_bpass_first", None)
        cs    = self._am_centroids[0] if self._am_centroids else []
        thr   = getattr(self, "_am_threshold", None)
        cal   = getattr(self, "_am_calibration_peaks", None)
        nseq  = getattr(self, "_am_per_frame_n", None)
        bad   = getattr(self, "_am_bad_frames", []) or []

        win = FigureWindow("Am-241 detection diagnostics",
                           size=(960, 620))
        win.figure.clf()
        ax1 = win.figure.add_subplot(2, 2, 1)
        ax2 = win.figure.add_subplot(2, 2, 2)
        ax3 = win.figure.add_subplot(2, 1, 2)

        ax1.imshow(np.asarray(frame, dtype=float), cmap="gray")
        ax1.set_title("Frame 1: raw + detected centroids")
        ax2.imshow(np.asarray(bp, dtype=float) if bp is not None
                    else np.asarray(frame, dtype=float),
                    cmap="magma")
        title2 = "Frame 1: bandpass"
        if thr is not None:
            title2 += f"  (threshold = {thr:.2f})"
        ax2.set_title(title2)
        for (x, y) in cs:
            for a in (ax1, ax2):
                a.plot(x, y, "o", markersize=14, markerfacecolor="none",
                       markeredgecolor=(0, 1, 0), markeredgewidth=2)
        for a in (ax1, ax2):
            a.set_xticks([]); a.set_yticks([])

        if nseq is not None:
            xs_n = np.arange(1, len(nseq) + 1)
            colors = ["#1f77b4" if n == 2 else "#d62728" for n in nseq]
            ax3.bar(xs_n, nseq, color=colors)
            ax3.axhline(2, color="black", linestyle="--", linewidth=1)
            ax3.set_xlabel("Am-241 frame index (1-based)")
            ax3.set_ylabel("# blobs above threshold")
            n_bad = len(bad)
            n_total = len(nseq)
            ax3.set_title(
                f"Per-frame fiducial count   "
                f"(good = {n_total - n_bad} / {n_total}, "
                f"rejected = {n_bad})")
            ax3.set_ylim(0, max(4, max(nseq) + 1))
        else:
            ax3.text(0.5, 0.5, "(no per-frame counts available)",
                     ha="center", va="center")
            ax3.set_xticks([]); ax3.set_yticks([])

        if cal is not None:
            win.figure.text(
                0.01, 0.97,
                f"Calibration peak intensities: {cal[0]:.2f}, {cal[1]:.2f}    "
                f"Threshold = 0.5 × min(peaks) = {thr:.2f}",
                fontsize=9, color="dimgray")

        win.canvas.draw_idle()
        win.show()
    def GetMCCButtonPushed(self):                      # MATLAB 856
        """Port of MATLAB GetMCCButtonPushed (line 856).

        Implements the core MCC analysis path - retention curves, clearance
        and difference maps, AUC60/90, fast/slow pixel ratios. Does NOT
        implement: Am-241 fiducial transforms, per-frame manual ROI mode,
        the .avi MCC movie, or the ShowScan frame-by-frame viewer. The
        corresponding checkboxes are currently treated as off.
        """
        try:
            if ndi is None:
                raise RuntimeError(
                    "scipy is required for Get MCC. pip install scipy")
            if self.MCCstack is None:
                raise RuntimeError(
                    "No scan stack loaded. Use Browse next to Scan Stack first.")
            if self.RLmask is None or self.Cmask is None or self.Pmask is None:
                raise RuntimeError(
                    "Run Analyze Depo first - it computes the central and "
                    "peripheral lung masks that MCC needs.")
            if self.BKGim is None:
                raise RuntimeError("BKG image is required.")
            if self.HotColdData is None:
                # Fall back to a length-7 NaN buffer so we can still write
                # FR / SLR into slots [5:7].
                self.HotColdData = np.full(7, np.nan, dtype=float)
            elif len(self.HotColdData) < 7:
                hcd = np.full(7, np.nan, dtype=float)
                hcd[: len(self.HotColdData)] = self.HotColdData
                self.HotColdData = hcd

            manual_mode = self.ManuallyCheckBox.isChecked()
            manual_cancelled = False

            stack = self.MCCstack          # shape (n_frames, H, W, C)
            n_frames = int(stack.shape[0])
            t = self._build_mcc_time_vector(n_frames)
            if t.size != n_frames:
                # Trim/pad to match.
                t = np.resize(t, n_frames)
            am_transforms = None
            am_good_idxs = list(range(1, n_frames + 1))
            if self.Am241CheckBox.isChecked():
                try:
                    am_transforms, am_good_idxs = self._get_americium_transforms()
                    self._show_am_debug_popup()
                except Exception as exc:
                    QMessageBox.warning(
                        self, "Am-241 tracking failed",
                        "Could not compute Am-241 transforms; falling back "
                        "to identity. Reason:" + chr(10) + str(exc))
                    am_transforms = None
                    am_good_idxs = list(range(1, n_frames + 1))

            HALF_LIFE = np.log(2.0) / (6.04 * 60.0)
            bkg_arr = np.asarray(self.BKGim, dtype=float)
            with np.errstate(invalid="ignore"):
                bkg_c = float(np.nanmedian(np.nanmedian(bkg_arr, axis=0)))

            rl = self.RLmask.astype(bool)
            cm = self.Cmask.astype(bool)
            pm = self.Pmask.astype(bool)

            wl_counts = np.zeros(n_frames)
            cr_counts = np.zeros(n_frames)
            pr_counts = np.zeros(n_frames)

            mcc_im0 = mcc_im60 = mcc_im90 = None
            idx_60 = int(np.argmin(np.abs(t - 60.0))) if (t == 60.0).any() else None
            idx_90 = int(np.argmin(np.abs(t - 90.0))) if (t == 90.0).any() else None

            cur_rl = rl.copy()
            cur_cm = cm.copy()
            cur_pm = pm.copy()
            og_verts = (np.asarray(self.ROIposition, dtype=float)
                        if self.ROIposition is not None else None)

            for i in range(n_frames):
                # MATLAB precedence: i==1 OR (not Am241 and not Manual) -> use
                # the static deposition masks; manual mode overrides; else
                # use the Am-241 transforms (warps RLmask only - CR/PR stay).
                if manual_mode and i > 0 and not manual_cancelled and og_verts is not None:
                    fr_disp = self._scan_stack_frame_2d(stack, i) - bkg_c
                    dlg = ManualROIDialog(
                        fr_disp, og_verts, i, n_frames, parent=self)
                    new_verts, cancelled = dlg.exec_modal()
                    if cancelled:
                        manual_cancelled = True
                    elif new_verts is not None:
                        M = self._affine_from_points(og_verts, new_verts)
                        cur_rl = self._warp_mask(rl, M)
                        cur_cm = self._warp_mask(cm, M)
                        cur_pm = self._warp_mask(pm, M)
                elif (am_transforms is not None and i > 0
                      and i - 1 < len(am_transforms)):
                    # Am-241: warp ONLY RLmask. MATLAB leaves CR/PR static.
                    M = am_transforms[i - 1]
                    cur_rl = self._warp_mask(rl, M)
                    # cur_cm, cur_pm stay at their previous values.

                fr = self._scan_stack_frame_2d(stack, i)
                fr_blur = ndi.gaussian_filter(fr, sigma=2)
                wl_im = fr_blur * cur_rl
                cr_im = fr_blur * cur_cm
                pr_im = fr_blur * cur_pm

                base = float(wl_im.sum()) - bkg_c
                base_c = float(cr_im.sum()) - bkg_c
                base_p = float(pr_im.sum()) - bkg_c
                decay = np.exp(-HALF_LIFE * float(t[i]))
                wl_counts[i] = base   / decay if decay != 0 else np.nan
                cr_counts[i] = base_c / decay if decay != 0 else np.nan
                pr_counts[i] = base_p / decay if decay != 0 else np.nan

                if i == 0:
                    mcc_im0 = wl_im.copy()
                if idx_60 is not None and i == idx_60:
                    base = ndi.gaussian_filter(fr, sigma=2)
                    if (am_transforms is not None
                            and i - 1 < len(am_transforms)):
                        Mi = self._invert_2x3(am_transforms[i - 1])
                        base = self._warp_image(base, Mi, output_shape=base.shape)
                    mcc_im60 = base
                if idx_90 is not None and i == idx_90:
                    base = ndi.gaussian_filter(fr, sigma=2)
                    if (am_transforms is not None
                            and i - 1 < len(am_transforms)):
                        Mi = self._invert_2x3(am_transforms[i - 1])
                        base = self._warp_image(base, Mi, output_shape=base.shape)
                    mcc_im90 = base

            # ---- Retention curve --------------------------------------
            denom = wl_counts[0] if wl_counts[0] != 0 else np.nan
            ret_array = wl_counts / denom

            ret_win = FigureWindow("Retention Curve", size=(700, 520))
            # Filter to "good" frames (those where Am-241 found 2 centroids).
            # am_good_idxs is 1-based to match MATLAB; convert to 0-based.
            good0 = [g - 1 for g in am_good_idxs if 0 <= g - 1 < n_frames]
            t_plot   = t[good0]   if good0 else t
            ret_plot = ret_array[good0] if good0 else ret_array
            ret_win.ax.plot(t_plot, ret_plot, linewidth=2)
            if len(good0) < n_frames:
                ret_win.ax.set_title(
                    f"Retention Curve  ({len(good0)}/{n_frames} good frames)")
            ret_win.ax.set_xlabel("Time (min)")
            ret_win.ax.set_ylabel("Retention")
            ret_win.ax.set_ylim(0, 1.2)
            ret_win.ax.set_title("Retention Curve")
            ret_win.ax.grid(True, alpha=0.3)
            ret_win.canvas.draw_idle()
            ret_win.show()

            # ---- Fast / slow pixel ratios -----------------------------
            depo_rl = (np.asarray(mcc_im0, dtype=float) * rl).astype(float)
            if mcc_im60 is not None:
                t60_im = mcc_im60.astype(float) * rl
                dc60 = t60_im / np.exp(-HALF_LIFE * 60.0)

                # MCC histogram in pixels (matches MATLAB mccHist).
                with np.errstate(invalid="ignore", divide="ignore"):
                    mcc_hist = 1.0 - np.where(depo_rl != 0, dc60 / depo_rl, np.nan)
                mcc_hist[mcc_hist < -0.2] = np.nan

                hist_win = FigureWindow("MCC histogram", size=(640, 480))
                hist_win.ax.hist(mcc_hist[np.isfinite(mcc_hist)].ravel(), bins=30)
                hist_win.ax.set_xlabel("Clearance in pixel")
                hist_win.ax.set_ylabel("Frequency")
                hist_win.canvas.draw_idle()
                hist_win.show()

                # Fast/slow ratio on non-cold pixels.
                cold_mask = (self.ColdMask if self.ColdMask is not None
                             else np.zeros_like(rl)).astype(bool)
                non_cold = rl & ~cold_mask
                depo_in = depo_rl[non_cold]
                dc60_in = dc60[non_cold]
                with np.errstate(invalid="ignore", divide="ignore"):
                    px_mcc = 1.0 - np.where(depo_in != 0, dc60_in / depo_in, np.nan)
                px_mcc = px_mcc[np.isfinite(px_mcc)]
                if px_mcc.size:
                    fr_ratio  = float(np.sum(px_mcc > 0.30)) / float(px_mcc.size)
                    slr_ratio = float(np.sum(px_mcc < 0.05)) / float(px_mcc.size)
                else:
                    fr_ratio = slr_ratio = np.nan
                self.HotColdData[5] = fr_ratio
                self.HotColdData[6] = slr_ratio

                # Threshold-of-Clearance arrays (FastSlowArray).
                if px_mcc.size:
                    thresholds = np.arange(1, 21) / 20.0
                    toc = np.array([(px_mcc > th).mean() for th in thresholds])
                    self.FastSlowArray = np.vstack([thresholds, toc])
                else:
                    self.FastSlowArray = None

                # Clearance map (fast/slow map).
                with np.errstate(invalid="ignore", divide="ignore"):
                    fs_map = np.where(depo_rl != 0,
                                      (depo_rl - dc60) / depo_rl, 0.0)
                fs_map = fs_map * (~cold_mask)
                fs_map = fs_map * rl - 1000.0 * (~rl)
                fs_map[~np.isfinite(fs_map)] = 0
                fs_map[fs_map < 0] = 0

                fs_win = FigureWindow("Clearance Map", size=(720, 620))
                im = fs_win.ax.imshow(fs_map, cmap="jet")
                fs_win.ax.set_xticks([]); fs_win.ax.set_yticks([])
                fs_win.figure.colorbar(im, ax=fs_win.ax)
                fs_win.ax.set_title("Clearance Map")
                fs_win.canvas.draw_idle()
                fs_win.show()

                # Difference map.
                diff_map = (depo_rl - dc60) * (~cold_mask) * rl
                diff_map[~np.isfinite(diff_map)] = 0
                diff_map[diff_map < 0] = 0
                diff_win = FigureWindow("Difference Map", size=(720, 620))
                im = diff_win.ax.imshow(diff_map, cmap="jet")
                diff_win.ax.set_xticks([]); diff_win.ax.set_yticks([])
                diff_win.figure.colorbar(im, ax=diff_win.ax)
                diff_win.ax.set_title("Difference Map")
                diff_win.canvas.draw_idle()
                diff_win.show()
            else:
                QMessageBox.information(
                    self, "No 60-min frame",
                    "No frame at t=60 found - clearance maps and fast/slow "
                    "ratios will be skipped.")

            # ---- ToC at 90 minutes -----------------------------------
            if mcc_im90 is not None and self.ColdMask is not None:
                t90_im = mcc_im90.astype(float) * rl
                dc90 = t90_im / np.exp(-HALF_LIFE * 90.0)
                non_cold = rl & ~self.ColdMask.astype(bool)
                depo_in = depo_rl[non_cold]
                dc90_in = dc90[non_cold]
                with np.errstate(invalid="ignore", divide="ignore"):
                    px_mcc90 = 1.0 - np.where(depo_in != 0, dc90_in / depo_in, np.nan)
                px_mcc90 = px_mcc90[np.isfinite(px_mcc90)]
                if px_mcc90.size:
                    thr = np.arange(1, 21) / 20.0
                    self.FastSlowArray90 = np.array(
                        [(px_mcc90 > th).mean() for th in thr])

            # ---- MCC array ------------------------------------------
            cret  = cr_counts / (cr_counts[0] if cr_counts[0] != 0 else np.nan)
            pret  = pr_counts / (pr_counts[0] if pr_counts[0] != 0 else np.nan)
            self.MCCarray = np.column_stack(
                [t,
                 100.0 * (1.0 - ret_array),
                 100.0 * (1.0 - cret),
                 100.0 * (1.0 - pret)])

            # ---- AUC60 / AUC90 ---------------------------------------
            sl = ret_array[5::5]
            self.AUC60 = float(np.nanmean(sl[:6])) if sl.size >= 6 else np.nan
            self.AUC90 = float(np.nanmean(sl[:9])) if sl.size >= 9 else np.nan

            # ---- Summary popup ---------------------------------------
            # ---- ShowScan viewer (optional) ---------------------------
            if self.ShowScanCheckBox.isChecked() and self.ROIposition is not None:
                am_cs = getattr(self, "_am_centroids", None)
                viewer = ScanStackViewer(
                    stack, self.ROIposition,
                    bkg_median=bkg_c, times=t, cmap="viridis",
                    am_transforms=am_transforms,
                    am_centroids=am_cs)
                _popup_windows.append(viewer)
                viewer.destroyed.connect(
                    lambda *_: _popup_windows.remove(viewer)
                    if viewer in _popup_windows else None)
                viewer.show()

            self.MCCdone = True
            msg_lines = [
                f"Frames analyzed:   {n_frames}",
                f"Time span:         {t.min():.0f} - {t.max():.0f} min",
                f"Retention at end:  {ret_array[-1]:.3f}",
                f"AUC60:             {self.AUC60:.4f}",
                f"AUC90:             {self.AUC90:.4f}",
                f"Fast fraction (>0.30 clearance):   {self.HotColdData[5]:.3f}",
                f"Slow fraction (<0.05 clearance):   {self.HotColdData[6]:.3f}",
            ]
            QMessageBox.information(self, "MCC analysis", chr(10).join(msg_lines))

        except Exception as e:
            import traceback
            QMessageBox.critical(self, "Get MCC error",
                                 f"{e}" + chr(10) + chr(10) + traceback.format_exc())

    # ---- Export Data ------------------------------------------------
    @staticmethod
    def _to_uint8_image(arr):
        """Convert any float / bool / int array into a uint8 image (0..255)."""
        a = np.asarray(arr)
        if a.dtype == bool:
            return (a.astype(np.uint8) * 255)
        a = a.astype(float)
        finite = a[np.isfinite(a)]
        if finite.size == 0:
            return np.zeros(a.shape, dtype=np.uint8)
        lo, hi = float(finite.min()), float(finite.max())
        if hi - lo < 1e-12:
            return np.zeros(a.shape, dtype=np.uint8)
        a = np.where(np.isfinite(a), (a - lo) / (hi - lo), 0)
        return np.clip(a * 255, 0, 255).astype(np.uint8)

    def ExportDataButtonPushed(self):                  # MATLAB 1212
        """Export masks, polygon ROI, and a multi-section .xlsx of results."""
        try:
            try:
                import openpyxl
                from openpyxl.utils import get_column_letter
            except ImportError:
                raise RuntimeError(
                    "openpyxl is required to export to .xlsx.\n"
                    "Run:  pip install openpyxl")
            try:
                from PIL import Image
            except ImportError:
                raise RuntimeError(
                    "Pillow is required to write TIFF masks.\n"
                    "Run:  pip install Pillow")
            try:
                import scipy.io as sio
            except ImportError:
                sio = None

            if not (self.HCdone or self.MCCdone):
                QMessageBox.warning(
                    self, "Nothing to export",
                    "Run Analyze Depo and/or Get MCC before exporting.")
                return

            # Pick a destination folder. Default to SubDir, else prompt.
            sub_dir = self.SubDir
            if not sub_dir or not Path(sub_dir).is_dir():
                sub_dir = QFileDialog.getExistingDirectory(
                    self, "Choose export folder",
                    str(Path.home()))
                if not sub_dir:
                    return
                self.SubDir = sub_dir
            sub_dir = Path(sub_dir)

            # ---- TIFF masks --------------------------------------------
            written = []
            def _save_tiff(arr, name):
                if arr is None:
                    return
                img = Image.fromarray(self._to_uint8_image(arr))
                path = sub_dir / name
                img.save(path, format="TIFF")
                written.append(str(path))

            _save_tiff(self.HotMask,        "hot_mask.tif")
            _save_tiff(self.ColdMask,       "cold_mask.tif")
            _save_tiff(self.RLmask,         "RL_mask.tif")
            _save_tiff(self.maskedDepoIm,   "depo_masked.tif")

            # ---- patient_roi.mat (plain Nx2 vertices) ------------------
            if self.ROIposition is not None and sio is not None:
                verts = np.asarray(self.ROIposition, dtype=float)
                roi_path = sub_dir / "patient_roi.mat"
                sio.savemat(
                    str(roi_path),
                    {"lung_mask": {"Position": verts},
                     "vertices":  verts},
                    do_compression=False, oned_as="row")
                written.append(str(roi_path))

            # ---- ScanData_wAUC_and_CP_masks.xlsx -----------------------
            xlsx_path = sub_dir / "ScanData_wAUC_and_CP_masks_py.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "ScanData"

            HC_names  = ["Hot NR", "Cold NR", "Hot SR", "Skew",
                         "C/P", "FR", "SLR"]
            MCC_names = ["Time (min)", "% Clearance",
                         "% Central Clearance", "% Peripheral Clearance"]
            Thr_names = ["MCC Threshold",
                         "Fraction Above (t=60)",
                         "Fraction Above (t=90)"]
            AUC_names = ["AUC60", "AUC90"]

            def _put(cell, value):
                if value is None:
                    return
                if isinstance(value, (np.floating, np.integer)):
                    value = value.item()
                if isinstance(value, float) and not np.isfinite(value):
                    return
                ws[cell] = value

            def _row(start_col, row, values):
                for i, v in enumerate(values):
                    col = get_column_letter(start_col + i)
                    _put(f"{col}{row}", v)

            _row(1, 1, HC_names)
            if self.HotColdData is not None:
                _row(1, 2, list(np.asarray(self.HotColdData)))

            _row(1, 4, MCC_names)
            if self.MCCarray is not None:
                arr = np.asarray(self.MCCarray)
                for r, vals in enumerate(arr):
                    _row(1, 5 + r, list(vals))

            _put("I1", Thr_names[0])
            _put("I2", Thr_names[1])
            _put("I3", Thr_names[2])
            if self.FastSlowArray is not None:
                arr = np.asarray(self.FastSlowArray)
                _row(10, 1, list(arr[0]))
                _row(10, 2, list(arr[1]))
            if self.FastSlowArray90 is not None:
                _row(10, 3, list(np.asarray(self.FastSlowArray90)))

            _put("I5", AUC_names[0]); _put("J5", AUC_names[1])
            _put("I6", self.AUC60);   _put("J6", self.AUC90)

            wb.save(str(xlsx_path))
            written.append(str(xlsx_path))

            QMessageBox.information(
                self, "Export complete",
                "Wrote " + str(len(written)) + " file(s) to:" + chr(10)
                + str(sub_dir) + chr(10) + chr(10)
                + chr(10).join(Path(p).name for p in written))

        except Exception as e:
            import traceback
            QMessageBox.critical(
                self, "Export error",
                f"{e}" + chr(10) + chr(10) + traceback.format_exc())

    # ---- Stubs - port bodies later from extracted_matlab_code.m ------
    def BKGImageEditFieldValueChanged(self): pass         # MATLAB 478
    def BKGImageEditFieldValueChanging(self, _t): pass    # MATLAB 514
    def TxImageEditFieldValueChanged(self): pass          # MATLAB 534
    def TxImageEditFieldValueChanging(self, _t): pass     # MATLAB 570
    def ScanStackEditFieldValueChanged(self): pass        # MATLAB 575
    def ScanStackEditFieldValueChanging(self, _t): pass   # MATLAB 611
    def BlurButtonPushed(self): pass                      # MATLAB 616
    def FilterWidthEditFieldValueChanged(self, _v): pass  # MATLAB 635
    def UndoBlurButtonPushed(self): pass                  # MATLAB 640
    def ThresholdSliderValueChanging(self, _v): pass      # MATLAB 651
    def DilateandFillButtonPushed(self): pass             # MATLAB 659
    def GetsButtonPushed(self): pass                      # MATLAB 669
    # AnalyzeDepoButtonPushed is implemented above (MATLAB 749).
    # GetMCCButtonPushed implemented above (MATLAB 856).
    def ShowScanCheckBoxValueChanged(self, _s): pass      # MATLAB 1156
    # ExportDataButtonPushed implemented above (MATLAB 1212).

    def ManuallyCheckBoxValueChanged(self, _s):           # MATLAB 1250
        """Manually + Am-241 are mutually exclusive (per MATLAB)."""
        if self.ManuallyCheckBox.isChecked():
            self.Am241CheckBox.setChecked(False)
            self.ShowScanCheckBox.setChecked(False)

    def Am241CheckBoxValueChanged(self, _s):              # MATLAB 1259
        if self.Am241CheckBox.isChecked():
            self.ManuallyCheckBox.setChecked(False)

    def LungROIFileEditFieldValueChanged(self): pass      # MATLAB 1267
    def LungROIFileEditFieldValueChanging(self, _t): pass # MATLAB 1273


def _global_excepthook(exc_type, exc_value, exc_tb):
    """Show any uncaught exception in a dialog instead of crashing silently."""
    import traceback
    msg = "".join(traceback.format_exception(exc_type, exc_value, exc_tb))
    print(msg, file=sys.stderr)
    try:
        from PySide6.QtWidgets import QApplication, QMessageBox
        if QApplication.instance() is not None:
            QMessageBox.critical(None, "Unhandled error", msg)
    except Exception:
        pass


def main():
    sys.excepthook = _global_excepthook
    app = QApplication(sys.argv)
    if "MCC_UI_SCALE" not in os.environ:
        try:
            dpi = app.primaryScreen().logicalDotsPerInch()
            global UI_SCALE
            UI_SCALE = max(1.30, dpi / 96.0 * 1.30)
        except Exception:
            pass
    win = MCCHotColdGUI()
    win.show()
    return app.exec()


if __name__ == "__main__":
    sys.exit(main())
